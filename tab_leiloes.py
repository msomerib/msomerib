import tkinter as tk
from tkinter import ttk, messagebox
from database import get_conn
from widgets import tv_padrao, btn, FormDlg, AZUL, CINZA, BRANCO, VERMELHO, VERDE, LARANJA
from utils import fmt_moeda, fmt_data, parse_moeda, parse_data, STATUS_LEILAO, STATUS_PARCELA_DOC


class TabLeiloes(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent, padding=12)
        self._build()
        self.carregar_leiloes()

    def _build(self):
        self.paned = tk.PanedWindow(self, orient="horizontal", sashwidth=6, bg="#ccc")
        self.paned.pack(fill="both", expand=True)

        esq = ttk.Frame(self.paned, padding=4)
        self.paned.add(esq, minsize=260)
        ttk.Label(esq, text="Leiloes (ALN)", style="Header.TLabel").pack(anchor="w", pady=(0, 6))
        btn(esq, "+ Novo Leilao", self.novo_leilao, "Accent.TButton").pack(fill="x", pady=2)
        btn(esq, "Editar Leilao", self.editar_leilao).pack(fill="x", pady=2)
        btn(esq, "Excluir", self.excluir_leilao, "Danger.TButton").pack(fill="x", pady=2)
        btn(esq, "Importar Lotes (XLSX)", self.importar_lotes).pack(fill="x", pady=2)
        cols_l = ["Numero", "Descricao", "Status", "Qtd"]
        largs_l = [80, 160, 100, 50]
        self.tv_leiloes, frm_l = tv_padrao(esq, cols_l, largs_l, altura=20)
        frm_l.pack(fill="both", expand=True, pady=(6, 0))
        self.tv_leiloes.bind("<<TreeviewSelect>>", lambda _: self.carregar_lotes())

        dir_ = ttk.Frame(self.paned, padding=4)
        self.paned.add(dir_, minsize=600)
        ttk.Label(dir_, text="Lotes do Leilao", style="Header.TLabel").pack(anchor="w", pady=(0, 6))
        acoes = ttk.Frame(dir_)
        acoes.pack(fill="x", pady=(0, 4))
        btn(acoes, "+ Novo Lote", self.novo_lote, "Accent.TButton").pack(side="left", padx=3)
        btn(acoes, "Editar", self.editar_lote).pack(side="left", padx=3)
        btn(acoes, "Excluir", self.excluir_lote, "Danger.TButton").pack(side="left", padx=3)
        cols_lt = ["Lote", "Q", "L", "Matric.", "Endereco", "Nr",
                   "Arrematante", "Modo Pag.", "Vlr.Minimo", "Vlr.Arrematado",
                   "Ag.", "5%Gar.", "Pago5%", "Complementar",
                   "Termo", "Docs", "Res.", "Leil.", "Status"]
        largs_lt = [50, 45, 45, 75, 160, 45, 160, 100, 100, 110,
                    60, 80, 80, 110, 60, 60, 60, 60, 90]
        self.tv_lotes, frm_lt = tv_padrao(dir_, cols_lt, largs_lt, altura=20)
        frm_lt.pack(fill="both", expand=True)
        self.tv_lotes.bind("<Double-1>", lambda _: self.editar_lote())
        self.v_total = tk.StringVar()
        ttk.Label(dir_, textvariable=self.v_total, foreground="#666").pack(anchor="w", pady=2)

    def carregar_leiloes(self):
        for i in self.tv_leiloes.get_children(): self.tv_leiloes.delete(i)
        with get_conn() as conn:
            rows = conn.execute("SELECT * FROM leiloes ORDER BY numero DESC").fetchall()
        for r in rows:
            self.tv_leiloes.insert("", "end", iid=r["id"], values=(
                r["numero"], r["descricao"] or "", r["status"], r["qtd_imoveis"] or 0))

    def carregar_lotes(self):
        for i in self.tv_lotes.get_children(): self.tv_lotes.delete(i)
        sel = self.tv_leiloes.selection()
        if not sel: return
        lid = int(sel[0])
        with get_conn() as conn:
            rows = conn.execute("SELECT * FROM lotes_leilao WHERE leilao_id=? ORDER BY lote_leilao", (lid,)).fetchall()
        self.v_total.set(f"{len(rows)} lote(s)")
        for r in rows:
            tag = "pago" if r["status"] == "PAGO" else ("alerta" if r["status"] == "PENDENTE" else "")
            self.tv_lotes.insert("", "end", iid=r["id"], values=(
                r["lote_leilao"] or "", r["quadra"] or "", r["lote"] or "",
                r["matricula"] or "", r["endereco"] or "", r["numero"] or "",
                r["arrematante"] or "", r["modo_pagamento"] or "",
                fmt_moeda(r["valor_minimo"]) if r["valor_minimo"] else "",
                fmt_moeda(r["valor_arrematado"]) if r["valor_arrematado"] else "",
                fmt_moeda(r["agio"]) if r["agio"] else "",
                fmt_moeda(r["garantia_5pct"]) if r["garantia_5pct"] else "",
                fmt_moeda(r["valor_pago_5pct"]) if r["valor_pago_5pct"] else "",
                fmt_moeda(r["valor_complementar"]) if r["valor_complementar"] else "",
                r["doc_termo_arr"] or "", r["doc_pessoais"] or "",
                r["doc_comprov_res"] or "", r["doc_leiloeiro"] or "", r["status"] or ""
            ), tags=(tag,))

    def _sel_leilao(self):
        sel = self.tv_leiloes.selection()
        if not sel: messagebox.showwarning("Atencao", "Selecione um leilao."); return None
        return int(sel[0])

    def _sel_lote(self):
        sel = self.tv_lotes.selection()
        if not sel: messagebox.showwarning("Atencao", "Selecione um lote."); return None
        return int(sel[0])

    def novo_leilao(self):
        d = LeilaoDlg(self); self.wait_window(d); self.carregar_leiloes()

    def editar_leilao(self):
        lid = self._sel_leilao()
        if lid is None: return
        with get_conn() as conn:
            row = conn.execute("SELECT * FROM leiloes WHERE id=?", (lid,)).fetchone()
        d = LeilaoDlg(self, row); self.wait_window(d); self.carregar_leiloes()

    def excluir_leilao(self):
        lid = self._sel_leilao()
        if lid is None: return
        if messagebox.askyesno("Confirmar", "Excluir leilao e todos os lotes?"):
            with get_conn() as conn:
                conn.execute("DELETE FROM lotes_leilao WHERE leilao_id=?", (lid,))
                conn.execute("DELETE FROM leiloes WHERE id=?", (lid,))
            self.carregar_leiloes(); self.carregar_lotes()

    def novo_lote(self):
        lid = self._sel_leilao()
        if lid is None: return
        d = LoteDlg(self, leilao_id=lid); self.wait_window(d); self.carregar_lotes()

    def editar_lote(self):
        ltid = self._sel_lote()
        if ltid is None: return
        with get_conn() as conn:
            row = conn.execute("SELECT * FROM lotes_leilao WHERE id=?", (ltid,)).fetchone()
        d = LoteDlg(self, row=row); self.wait_window(d); self.carregar_lotes()

    def excluir_lote(self):
        ltid = self._sel_lote()
        if ltid is None: return
        if messagebox.askyesno("Confirmar", "Excluir este lote?"):
            with get_conn() as conn:
                conn.execute("DELETE FROM lotes_leilao WHERE id=?", (ltid,))
            self.carregar_lotes()

    def importar_lotes(self):
        lid = self._sel_leilao()
        if lid is None: return
        from tkinter.filedialog import askopenfilename
        path = askopenfilename(title="Selecionar planilha de leilao",
                               filetypes=[("Excel", "*.xlsx *.xlsm"), ("Todos", "*.*")])
        if not path: return
        ImportarLotesDlg(self, path, lid)
        self.carregar_lotes(); self.carregar_leiloes()


class LeilaoDlg(FormDlg):
    def __init__(self, parent, row=None):
        super().__init__(parent, "Editar Leilao" if row else "Novo Leilao", w=560, h=380)
        self.row = row; self._build()
        if row: self._fill()

    def _build(self):
        c = self.corpo
        c.columnconfigure(0, weight=1); c.columnconfigure(1, weight=1)
        self.v_num, _ = self.campo(c, "Numero (ex: 001.26) *", 0, 0)
        self.v_status, _ = self.combo(c, "Status", STATUS_LEILAO, 0, 1)
        self.v_desc, _ = self.campo(c, "Descricao", 1, 0, cs=2, w=50)
        self.v_data, _ = self.campo(c, "Data do Edital (dd/mm/aaaa)", 2, 0)
        self.v_qtd, _ = self.campo(c, "Qtd. Imoveis", 2, 1)
        self.t_obs = self.texto(c, "Observacoes", 3, 0, cs=2, h=3)
        btn(self.rodape, "Salvar", self._salvar, "Accent.TButton").pack(side="right", padx=8, pady=6)
        btn(self.rodape, "Cancelar", self.destroy).pack(side="right", pady=6)

    def _fill(self):
        r = self.row
        self.v_num.set(r["numero"]); self.v_status.set(r["status"])
        self.v_desc.set(r["descricao"] or ""); self.v_data.set(fmt_data(r["data_edital"]))
        self.v_qtd.set(r["qtd_imoveis"] or "")
        self.t_obs.insert("1.0", r["observacoes"] or "")

    def _salvar(self):
        num = self.v_num.get().strip()
        if not num: messagebox.showerror("Erro", "Numero e obrigatorio."); return
        d = dict(numero=num, status=self.v_status.get(),
                 descricao=self.v_desc.get().strip() or None,
                 data_edital=parse_data(self.v_data.get()),
                 qtd_imoveis=self.v_qtd.get().strip() or None,
                 observacoes=self.t_obs.get("1.0", "end-1c").strip() or None)
        with get_conn() as conn:
            if self.row:
                conn.execute("""UPDATE leiloes SET numero=:numero, status=:status,
                    descricao=:descricao, data_edital=:data_edital,
                    qtd_imoveis=:qtd_imoveis, observacoes=:observacoes WHERE id=?""",
                    {**d, "?": self.row["id"]})
            else:
                conn.execute("""INSERT INTO leiloes (numero,status,descricao,data_edital,qtd_imoveis,observacoes)
                    VALUES (:numero,:status,:descricao,:data_edital,:qtd_imoveis,:observacoes)""", d)
        self.destroy()


class LoteDlg(FormDlg):
    def __init__(self, parent, row=None, leilao_id=None):
        super().__init__(parent, "Editar Lote" if row else "Novo Lote", w=700, h=660)
        self.row = row
        self.leilao_id = leilao_id or (row["leilao_id"] if row else None)
        self._build()
        if row: self._fill()

    def _build(self):
        c = self.corpo
        for i in range(4): c.columnconfigure(i, weight=1)
        self.v_lote, _ = self.campo(c, "Lote Leilao", 0, 0, w=10)
        self.v_qd, _ = self.campo(c, "Quadra", 0, 1, w=10)
        self.v_lt, _ = self.campo(c, "Lote", 0, 2, w=10)
        self.v_mat, _ = self.campo(c, "Matricula", 0, 3, w=10)
        self.v_end, _ = self.campo(c, "Endereco", 1, 0, cs=3, w=50)
        self.v_num, _ = self.campo(c, "Nr", 1, 3, w=10)
        self.v_arr, _ = self.campo(c, "Arrematante", 2, 0, cs=2, w=40)
        self.v_dep, _ = self.campo(c, "Depositante", 2, 2, cs=2, w=40)
        self.v_modo, _ = self.campo(c, "Modo Pagamento", 3, 0)
        self.v_cpf, _ = self.campo(c, "CPF/CNPJ", 3, 1)
        self.v_rg, _ = self.campo(c, "RG", 3, 2)
        self.v_tel, _ = self.campo(c, "Telefone", 3, 3)
        self.v_conj, _ = self.campo(c, "Conjuge/Socios", 4, 0, cs=2, w=40)
        self.v_cpf_conj, _ = self.campo(c, "CPF Conjuge", 4, 2)
        self.v_rg_conj, _ = self.campo(c, "RG Conjuge", 4, 3)
        self.v_vmin, _ = self.campo(c, "Valor Minimo (R$)", 5, 0)
        self.v_varr, _ = self.campo(c, "Valor Arrematado (R$)", 5, 1)
        self.v_agio, _ = self.campo(c, "Agio (R$)", 5, 2)
        self.v_status, _ = self.combo(c, "Status", ["PENDENTE", "PAGO", "CANCELADO"], 5, 3)
        self.v_gar, _ = self.campo(c, "5% Garantia (R$)", 6, 0)
        self.v_pago5, _ = self.campo(c, "Pago 5% (R$)", 6, 1)
        self.v_compl, _ = self.campo(c, "Complementar (R$)", 6, 2)
        op = STATUS_PARCELA_DOC
        self.v_dtermo, _ = self.combo(c, "Termo Arrematacao", op, 8, 0)
        self.v_ddocs, _ = self.combo(c, "Doc. Pessoais", op, 8, 1)
        self.v_dres, _ = self.combo(c, "Comprov. Residencia", op, 8, 2)
        self.v_dleil, _ = self.combo(c, "Comprov. Leiloeiro", op, 8, 3)
        self.t_obs = self.texto(c, "OBS", 9, 0, cs=4, h=2)
        btn(self.rodape, "Salvar", self._salvar, "Accent.TButton").pack(side="right", padx=8, pady=6)
        btn(self.rodape, "Cancelar", self.destroy).pack(side="right", pady=6)

    def _fill(self):
        r = self.row
        self.v_lote.set(r["lote_leilao"] or ""); self.v_qd.set(r["quadra"] or "")
        self.v_lt.set(r["lote"] or ""); self.v_mat.set(r["matricula"] or "")
        self.v_end.set(r["endereco"] or ""); self.v_num.set(r["numero"] or "")
        self.v_arr.set(r["arrematante"] or ""); self.v_dep.set(r["depositante"] or "")
        self.v_modo.set(r["modo_pagamento"] or ""); self.v_cpf.set(r["cpf_cnpj"] or "")
        self.v_rg.set(r["rg"] or ""); self.v_tel.set(r["telefone"] or "")
        self.v_conj.set(r["conjuge"] or ""); self.v_cpf_conj.set(r["cpf_conjuge"] or "")
        self.v_rg_conj.set(r["rg_conjuge"] or "")
        self.v_vmin.set(r["valor_minimo"] or ""); self.v_varr.set(r["valor_arrematado"] or "")
        self.v_agio.set(r["agio"] or ""); self.v_status.set(r["status"] or "PENDENTE")
        self.v_gar.set(r["garantia_5pct"] or ""); self.v_pago5.set(r["valor_pago_5pct"] or "")
        self.v_compl.set(r["valor_complementar"] or "")
        self.v_dtermo.set(r["doc_termo_arr"] or "NAO"); self.v_ddocs.set(r["doc_pessoais"] or "NAO")
        self.v_dres.set(r["doc_comprov_res"] or "NAO"); self.v_dleil.set(r["doc_leiloeiro"] or "NAO")
        self.t_obs.insert("1.0", r["observacoes"] or "")

    def _salvar(self):
        d = dict(
            leilao_id=self.leilao_id,
            lote_leilao=self.v_lote.get().strip() or None,
            quadra=self.v_qd.get().strip() or None, lote=self.v_lt.get().strip() or None,
            matricula=self.v_mat.get().strip() or None, endereco=self.v_end.get().strip() or None,
            numero=self.v_num.get().strip() or None, arrematante=self.v_arr.get().strip() or None,
            depositante=self.v_dep.get().strip() or None, modo_pagamento=self.v_modo.get().strip() or None,
            cpf_cnpj=self.v_cpf.get().strip() or None, rg=self.v_rg.get().strip() or None,
            telefone=self.v_tel.get().strip() or None, conjuge=self.v_conj.get().strip() or None,
            cpf_conjuge=self.v_cpf_conj.get().strip() or None, rg_conjuge=self.v_rg_conj.get().strip() or None,
            valor_minimo=parse_moeda(self.v_vmin.get()) or None,
            valor_arrematado=parse_moeda(self.v_varr.get()) or None,
            agio=parse_moeda(self.v_agio.get()) or None, status=self.v_status.get(),
            garantia_5pct=parse_moeda(self.v_gar.get()) or None,
            valor_pago_5pct=parse_moeda(self.v_pago5.get()) or None,
            valor_complementar=parse_moeda(self.v_compl.get()) or None,
            doc_termo_arr=self.v_dtermo.get(), doc_pessoais=self.v_ddocs.get(),
            doc_comprov_res=self.v_dres.get(), doc_leiloeiro=self.v_dleil.get(),
            observacoes=self.t_obs.get("1.0", "end-1c").strip() or None,
        )
        with get_conn() as conn:
            if self.row:
                conn.execute("""UPDATE lotes_leilao SET lote_leilao=:lote_leilao,
                    quadra=:quadra,lote=:lote,matricula=:matricula,endereco=:endereco,
                    numero=:numero,arrematante=:arrematante,depositante=:depositante,
                    modo_pagamento=:modo_pagamento,cpf_cnpj=:cpf_cnpj,rg=:rg,
                    telefone=:telefone,conjuge=:conjuge,cpf_conjuge=:cpf_conjuge,
                    rg_conjuge=:rg_conjuge,valor_minimo=:valor_minimo,
                    valor_arrematado=:valor_arrematado,agio=:agio,status=:status,
                    garantia_5pct=:garantia_5pct,valor_pago_5pct=:valor_pago_5pct,
                    valor_complementar=:valor_complementar,doc_termo_arr=:doc_termo_arr,
                    doc_pessoais=:doc_pessoais,doc_comprov_res=:doc_comprov_res,
                    doc_leiloeiro=:doc_leiloeiro,observacoes=:observacoes WHERE id=?""",
                    {**d, "?": self.row["id"]})
            else:
                conn.execute("""INSERT INTO lotes_leilao (leilao_id,lote_leilao,quadra,lote,
                    matricula,endereco,numero,arrematante,depositante,modo_pagamento,cpf_cnpj,
                    rg,telefone,conjuge,cpf_conjuge,rg_conjuge,valor_minimo,valor_arrematado,
                    agio,status,garantia_5pct,valor_pago_5pct,valor_complementar,
                    doc_termo_arr,doc_pessoais,doc_comprov_res,doc_leiloeiro,observacoes)
                    VALUES (:leilao_id,:lote_leilao,:quadra,:lote,:matricula,:endereco,
                    :numero,:arrematante,:depositante,:modo_pagamento,:cpf_cnpj,:rg,
                    :telefone,:conjuge,:cpf_conjuge,:rg_conjuge,:valor_minimo,
                    :valor_arrematado,:agio,:status,:garantia_5pct,:valor_pago_5pct,
                    :valor_complementar,:doc_termo_arr,:doc_pessoais,:doc_comprov_res,
                    :doc_leiloeiro,:observacoes)""", d)
        self.destroy()


class ImportarLotesDlg(tk.Toplevel):
    def __init__(self, parent, path, leilao_id):
        super().__init__(parent)
        self.title("Importar Lotes do Leilao - XLSX")
        self.geometry("680x460")
        self.grab_set()
        self.path = path; self.leilao_id = leilao_id
        self._build()
        self.after(100, self._carregar_abas)

    def _build(self):
        ttk.Label(self, text=f"Arquivo: {self.path.split('/')[-1]}",
                  font=("Segoe UI", 10)).pack(anchor="w", padx=12, pady=8)
        f = ttk.Frame(self); f.pack(fill="x", padx=12)
        ttk.Label(f, text="Aba:").pack(side="left")
        self.v_aba = tk.StringVar()
        self.cb_aba = ttk.Combobox(f, textvariable=self.v_aba, state="readonly", width=35)
        self.cb_aba.pack(side="left", padx=6)
        self.log = tk.Text(self, height=17, font=("Consolas", 9))
        self.log.pack(fill="both", expand=True, padx=12, pady=4)
        rodape = ttk.Frame(self); rodape.pack(fill="x", padx=12, pady=6)
        btn(rodape, "Importar", self._importar, "Accent.TButton").pack(side="right", padx=4)
        btn(rodape, "Fechar", self.destroy).pack(side="right")

    def _log(self, t): self.log.insert("end", t + "\n"); self.log.see("end")

    def _carregar_abas(self):
        import openpyxl
        wb = openpyxl.load_workbook(self.path, read_only=True, data_only=True, keep_vba=False)
        self.cb_aba["values"] = wb.sheetnames
        if wb.sheetnames: self.cb_aba.set(wb.sheetnames[0])
        wb.close()

    def _importar(self):
        import openpyxl
        aba = self.v_aba.get()
        wb = openpyxl.load_workbook(self.path, read_only=True, data_only=True, keep_vba=False)
        ws = wb[aba]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        header = None; start = 0
        for i, row in enumerate(rows):
            vals = [str(v).upper().strip() if v else "" for v in row]
            if "ARREMATANTE" in vals or "LOTE LEILAO" in vals:
                header = vals; start = i + 1; break
        if header is None: self._log("Cabecalho nao encontrado."); return

        def ci(h, *nomes):
            for nm in nomes:
                for j, v in enumerate(h):
                    if nm in v: return j
            return None

        i_lote_l = ci(header, "LOTE LEILAO", "LOTE ")
        i_arr = ci(header, "ARREMATANTE")
        i_dep = ci(header, "DEPOSITANTE")
        i_modo = ci(header, "MODO PAG", "FORM")
        i_cpf = ci(header, "CPF")
        i_rg = ci(header, "RG")
        i_conj = ci(header, "CONJUGE")
        i_cpfconj = ci(header, "CPF CONJUGE")
        i_rgconj = ci(header, "RG CONJUGE")
        i_tel = ci(header, "TELEFONE")
        i_email = ci(header, "E-MAIL", "EMAIL")
        i_qd = ci(header, "QUADRA")
        i_lt = ci(header, "LOTE")
        i_mat = ci(header, "MATRICULA")
        i_end = ci(header, "ENDERECO")
        i_num = ci(header, "NR", "NUMERO")
        i_vmin = ci(header, "VALOR MINIMO")
        i_varr = ci(header, "VALOR ARREMATADO")
        i_agio = ci(header, "AGIO")
        i_gar = ci(header, "5% GARANTIA", "GARANTIA")
        i_pago5 = ci(header, "PAGO 5%")
        i_compl = ci(header, "COMPLEMENTAR")
        i_dtermo = ci(header, "TERMO")
        i_ddocs = ci(header, "DOC PESSOAIS")
        i_dres = ci(header, "COMPROV. RESID")
        i_dleil = ci(header, "COMPROV. LEILOEIRO")
        i_obs = ci(header, "OBS")

        ok = skip = 0
        with get_conn() as conn:
            for row in rows[start:]:
                if all(v is None for v in row): continue
                def g(i): return row[i] if i is not None and i < len(row) else None
                lote_l = g(i_lote_l)
                if lote_l is None: skip += 1; continue
                d = dict(
                    leilao_id=self.leilao_id, lote_leilao=lote_l,
                    arrematante=str(g(i_arr) or "").strip() or None,
                    depositante=str(g(i_dep) or "").strip() or None,
                    modo_pagamento=str(g(i_modo) or "").strip() or None,
                    cpf_cnpj=str(g(i_cpf) or "").strip() or None,
                    rg=str(g(i_rg) or "").strip() or None,
                    conjuge=str(g(i_conj) or "").strip() or None,
                    cpf_conjuge=str(g(i_cpfconj) or "").strip() or None,
                    rg_conjuge=str(g(i_rgconj) or "").strip() or None,
                    telefone=str(g(i_tel) or "").strip() or None,
                    email=str(g(i_email) or "").strip() or None,
                    quadra=g(i_qd), lote=g(i_lt), matricula=g(i_mat),
                    endereco=str(g(i_end) or "").strip() or None,
                    numero=str(g(i_num) or "").strip() or None,
                    valor_minimo=g(i_vmin), valor_arrematado=g(i_varr), agio=g(i_agio),
                    garantia_5pct=g(i_gar), valor_pago_5pct=g(i_pago5), valor_complementar=g(i_compl),
                    doc_termo_arr=str(g(i_dtermo) or "NAO"),
                    doc_pessoais=str(g(i_ddocs) or "NAO"),
                    doc_comprov_res=str(g(i_dres) or "NAO"),
                    doc_leiloeiro=str(g(i_dleil) or "NAO"),
                    observacoes=str(g(i_obs) or "").strip() or None,
                )
                try:
                    conn.execute("""INSERT INTO lotes_leilao (leilao_id,lote_leilao,arrematante,
                        depositante,modo_pagamento,cpf_cnpj,rg,conjuge,cpf_conjuge,rg_conjuge,
                        telefone,email,quadra,lote,matricula,endereco,numero,valor_minimo,
                        valor_arrematado,agio,garantia_5pct,valor_pago_5pct,valor_complementar,
                        doc_termo_arr,doc_pessoais,doc_comprov_res,doc_leiloeiro,observacoes)
                        VALUES (:leilao_id,:lote_leilao,:arrematante,:depositante,:modo_pagamento,
                        :cpf_cnpj,:rg,:conjuge,:cpf_conjuge,:rg_conjuge,:telefone,:email,
                        :quadra,:lote,:matricula,:endereco,:numero,:valor_minimo,:valor_arrematado,
                        :agio,:garantia_5pct,:valor_pago_5pct,:valor_complementar,:doc_termo_arr,
                        :doc_pessoais,:doc_comprov_res,:doc_leiloeiro,:observacoes)""", d)
                    ok += 1
                except Exception as e:
                    self._log(f"  Erro lote {lote_l}: {e}"); skip += 1
            conn.execute("UPDATE leiloes SET qtd_imoveis=(SELECT COUNT(*) FROM lotes_leilao WHERE leilao_id=?) WHERE id=?",
                         (self.leilao_id, self.leilao_id))
        self._log(f"\nConcluido: {ok} lotes inseridos, {skip} ignorados.")
