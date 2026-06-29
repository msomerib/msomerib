import tkinter as tk
from tkinter import ttk, messagebox
from database import get_conn
from widgets import AZUL, CINZA, BRANCO, VERDE, VERMELHO, LARANJA, AMARELO
from utils import fmt_moeda


class TabRelatorios(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent, padding=12)
        self._build()

    def _build(self):
        ttk.Label(self, text="Relatorios e Exportacoes",
                  style="Header.TLabel").pack(anchor="w", pady=(0, 12))
        nb = ttk.Notebook(self)
        nb.pack(fill="both", expand=True)

        f1 = ttk.Frame(nb, padding=12)
        nb.add(f1, text="  Verificacao Geral  ")
        self._build_verificacao(f1)

        f2 = ttk.Frame(nb, padding=12)
        nb.add(f2, text="  Contratos Vencidos  ")
        self._build_contratos_vencidos(f2)

        f3 = ttk.Frame(nb, padding=12)
        nb.add(f3, text="  Resumo Leiloes  ")
        self._build_leiloes(f3)

        f4 = ttk.Frame(nb, padding=12)
        nb.add(f4, text="  Exportar Excel  ")
        self._build_exportar(f4)

    def _build_verificacao(self, parent):
        ttk.Label(parent, text="Totais por Status dos Imoveis",
                  font=("Segoe UI", 11, "bold")).pack(anchor="w", pady=(0, 8))
        self.txt_ver = tk.Text(parent, height=20, font=("Consolas", 11), bg=BRANCO)
        self.txt_ver.pack(fill="both", expand=True)
        from widgets import btn
        btn(parent, "Atualizar", self._atualizar_verificacao).pack(anchor="e", pady=6)
        self._atualizar_verificacao()

    def _atualizar_verificacao(self):
        self.txt_ver.delete("1.0", "end")
        with get_conn() as conn:
            tot = conn.execute("SELECT COUNT(*) FROM imoveis").fetchone()[0]
            por_status = conn.execute("SELECT status, COUNT(*) as qt FROM imoveis GROUP BY status ORDER BY qt DESC").fetchall()
            por_tipo_ct = conn.execute("SELECT tipo_contrato, COUNT(*) as qt FROM imoveis WHERE tipo_contrato IS NOT NULL GROUP BY tipo_contrato ORDER BY qt DESC").fetchall()
            ct_vigente = conn.execute("SELECT COUNT(*) FROM contratos WHERE status='VIGENTE'").fetchone()[0]
            ct_vencido = conn.execute("SELECT COUNT(*) FROM contratos WHERE status='VENCIDO'").fetchone()[0]
            leiloes = conn.execute("SELECT COUNT(*) FROM leiloes").fetchone()[0]
            lotes = conn.execute("SELECT COUNT(*) FROM lotes_leilao").fetchone()[0]
            ofertas = conn.execute("SELECT COUNT(*) FROM ofertas_compra").fetchone()[0]
            pgto_ok = conn.execute("SELECT COUNT(*) FROM ofertas_compra WHERE status_valor_pgto='PGTO. FINALIZADO'").fetchone()[0]
            vl_total = conn.execute("SELECT SUM(valor_total) FROM ofertas_compra").fetchone()[0] or 0
            vl_pago = conn.execute("SELECT SUM(valor_pago) FROM ofertas_compra").fetchone()[0] or 0
        lines = ["=" * 50, f"  TOTAL DE IMOVEIS CADASTRADOS:  {tot}", "=" * 50, "", "  POR STATUS:"]
        for r in por_status:
            lines.append(f"    {r['status']:<20} {r['qt']:>5}")
        lines += ["", "  POR TIPO DE CONTRATO:"]
        for r in por_tipo_ct:
            lines.append(f"    {r['tipo_contrato']:<20} {r['qt']:>5}")
        lines += ["", "=" * 50,
            f"  CONTRATOS VIGENTES:  {ct_vigente}",
            f"  CONTRATOS VENCIDOS:  {ct_vencido}",
            "", f"  LEILOES CADASTRADOS: {leiloes}",
            f"  LOTES EM LEILOES:    {lotes}",
            "", f"  OFERTAS DE COMPRA:   {ofertas}",
            f"  PAGAMENTOS CONCLUIDOS: {pgto_ok}",
            "", f"  VALOR TOTAL OFERTAS:  {fmt_moeda(vl_total)}",
            f"  VALOR TOTAL PAGO:     {fmt_moeda(vl_pago)}",
            f"  SALDO A RECEBER:      {fmt_moeda(vl_total - vl_pago)}",
            "=" * 50]
        self.txt_ver.insert("1.0", "\n".join(lines))

    def _build_contratos_vencidos(self, parent):
        from widgets import tv_padrao, btn
        from utils import fmt_data
        ttk.Label(parent, text="Contratos Vencidos ou Vencendo em 90 dias",
                  font=("Segoe UI", 11, "bold")).pack(anchor="w")
        btn(parent, "Atualizar", lambda: self._carregar_vencidos()).pack(anchor="e", pady=4)
        cols = ["Nr Contrato", "Entidade", "Modalidade", "Vigencia", "Status", "Dias/Vencer",
                "Prorrog. ate", "Telefone", "E-mail"]
        largs = [130, 220, 100, 90, 80, 90, 100, 140, 200]
        self.tv_venc, frm = tv_padrao(parent, cols, largs, altura=18)
        frm.pack(fill="both", expand=True)
        self._carregar_vencidos()

    def _carregar_vencidos(self):
        from utils import fmt_data
        import datetime
        for i in self.tv_venc.get_children(): self.tv_venc.delete(i)
        limite = (datetime.date.today() + datetime.timedelta(days=90)).isoformat()
        with get_conn() as conn:
            rows = conn.execute("""SELECT *, julianday(data_vigencia) - julianday('now') as dias
                FROM contratos WHERE status IN ('VIGENTE','VENCIDO') AND data_vigencia <= ?
                ORDER BY data_vigencia""", (limite,)).fetchall()
        for r in rows:
            dias = int(r["dias"]) if r["dias"] is not None else ""
            tag = "vencido" if (isinstance(dias, int) and dias < 0) else ("alerta" if (isinstance(dias, int) and dias <= 30) else "")
            self.tv_venc.insert("", "end", values=(
                r["num_contrato"], r["entidade"][:35],
                r["modalidade"] or "", fmt_data(r["data_vigencia"]),
                r["status"], dias, fmt_data(r["prorrogavel_ate"]),
                r["telefone"] or "", r["email"] or ""
            ), tags=(tag,))

    def _build_leiloes(self, parent):
        from widgets import tv_padrao, btn
        ttk.Label(parent, text="Resumo por Leilao",
                  font=("Segoe UI", 11, "bold")).pack(anchor="w")
        btn(parent, "Atualizar", lambda: self._carregar_resumo_leiloes()).pack(anchor="e", pady=4)
        cols = ["Leilao", "Descricao", "Status", "Total Lotes",
                "Pago", "Pendente", "Cancelado", "Valor Minimo Total", "Valor Arrematado Total"]
        largs = [80, 200, 100, 80, 60, 70, 80, 140, 160]
        self.tv_leil, frm = tv_padrao(parent, cols, largs, altura=18)
        frm.pack(fill="both", expand=True)
        self._carregar_resumo_leiloes()

    def _carregar_resumo_leiloes(self):
        for i in self.tv_leil.get_children(): self.tv_leil.delete(i)
        with get_conn() as conn:
            leiloes = conn.execute("SELECT * FROM leiloes ORDER BY numero DESC").fetchall()
            for l in leiloes:
                stats = conn.execute("SELECT status, COUNT(*) as qt FROM lotes_leilao WHERE leilao_id=? GROUP BY status", (l["id"],)).fetchall()
                vmin = conn.execute("SELECT SUM(valor_minimo) FROM lotes_leilao WHERE leilao_id=?", (l["id"],)).fetchone()[0] or 0
                varr = conn.execute("SELECT SUM(valor_arrematado) FROM lotes_leilao WHERE leilao_id=?", (l["id"],)).fetchone()[0] or 0
                total = conn.execute("SELECT COUNT(*) FROM lotes_leilao WHERE leilao_id=?", (l["id"],)).fetchone()[0]
                st_map = {r["status"]: r["qt"] for r in stats}
                self.tv_leil.insert("", "end", values=(
                    l["numero"], l["descricao"] or "", l["status"], total,
                    st_map.get("PAGO", 0), st_map.get("PENDENTE", 0), st_map.get("CANCELADO", 0),
                    fmt_moeda(vmin), fmt_moeda(varr)))

    def _build_exportar(self, parent):
        ttk.Label(parent, text="Exportar dados para Excel (.xlsx)",
                  font=("Segoe UI", 11, "bold")).pack(anchor="w", pady=(0, 12))
        for titulo, cmd in [
            ("Exportar Imoveis", self._exp_imoveis),
            ("Exportar Contratos", self._exp_contratos),
            ("Exportar Lotes de Leiloes", self._exp_lotes),
            ("Exportar Ofertas de Compra", self._exp_ofertas),
        ]:
            f = tk.Frame(parent, bg=CINZA)
            f.pack(fill="x", pady=4)
            tk.Label(f, text=titulo, font=("Segoe UI", 11), bg=CINZA, width=30,
                     anchor="w").pack(side="left")
            from widgets import btn
            btn(f, "Exportar", cmd).pack(side="left", padx=8)
        self.lbl_status = ttk.Label(parent, text="")
        self.lbl_status.pack(anchor="w", pady=8)

    def _exportar(self, titulo, query, colunas):
        from tkinter.filedialog import asksaveasfilename
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        path = asksaveasfilename(defaultextension=".xlsx",
                                 filetypes=[("Excel", "*.xlsx")],
                                 title=f"Salvar {titulo}")
        if not path: return
        with get_conn() as conn:
            rows = conn.execute(query).fetchall()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = titulo[:30]
        azul_fill = PatternFill("solid", fgColor="1565C0")
        hdr_font = Font(bold=True, color="FFFFFF")
        for j, col in enumerate(colunas, 1):
            c = ws.cell(row=1, column=j, value=col)
            c.font = hdr_font; c.fill = azul_fill
            c.alignment = Alignment(horizontal="center")
            ws.column_dimensions[c.column_letter].width = max(12, len(col)+4)
        for i, row in enumerate(rows, 2):
            for j, val in enumerate(row, 1):
                ws.cell(row=i, column=j, value=val)
        wb.save(path)
        self.lbl_status.config(text=f"Exportado: {path.split('/')[-1]} ({len(rows)} registros)")
        messagebox.showinfo("Sucesso", f"Arquivo salvo:\n{path}")

    def _exp_imoveis(self):
        self._exportar("Imoveis",
            "SELECT codigo,entidade,morador,quadra,lote,endereco,numero,tipo,area_m2,construcao_m2,matricula,local_area,num_contrato,tipo_contrato,vigencia,sit_contrato,taxa_ocupacao,preco_minimo,preco_25pct,status,observacoes FROM imoveis ORDER BY codigo",
            ["Codigo","Entidade","Morador","Quadra","Lote","Endereco","Nr","Tipo","Area m2","Constr m2","Matricula","Local","Nr Contrato","Tipo CT","Vigencia","Sit. Contrato","Tx.Ocup.","Preco Min.","Preco 25%","Status","Observacoes"])

    def _exp_contratos(self):
        self._exportar("Contratos",
            "SELECT num_contrato,entidade,modalidade,tipo_uso,assinatura_dig,data_vigencia,status,dias_vencer,prorrogavel_ate,cpf_cnpj,responsavel,end_correspondencia,telefone,email,observacoes FROM contratos ORDER BY data_vigencia",
            ["Nr Contrato","Entidade","Modalidade","Tipo Uso","Ass.Dig.","Vigencia","Status","Dias Vencer","Prorrog. ate","CPF/CNPJ","Responsavel","End. Correspondencia","Telefone","E-mail","Observacoes"])

    def _exp_lotes(self):
        self._exportar("Lotes Leilao",
            """SELECT l.numero, lt.lote_leilao, lt.quadra, lt.lote, lt.matricula,
               lt.endereco, lt.numero, lt.arrematante, lt.depositante, lt.modo_pagamento,
               lt.cpf_cnpj, lt.rg, lt.telefone, lt.email,
               lt.valor_minimo, lt.valor_arrematado, lt.agio,
               lt.garantia_5pct, lt.valor_pago_5pct, lt.valor_complementar,
               lt.doc_termo_arr, lt.doc_pessoais, lt.doc_comprov_res, lt.doc_leiloeiro,
               lt.status, lt.observacoes
               FROM lotes_leilao lt JOIN leiloes l ON l.id=lt.leilao_id
               ORDER BY l.numero, lt.lote_leilao""",
            ["ALN","Lote Leilao","Quadra","Lote","Matricula","Endereco","Nr","Arrematante",
             "Depositante","Modo Pgto","CPF/CNPJ","RG","Telefone","E-mail",
             "Vlr.Minimo","Vlr.Arrematado","Agio","5%Gar.","Pago5%","Complementar",
             "Termo","Docs","Resid.","Leiloeiro","Status","Obs"])

    def _exp_ofertas(self):
        self._exportar("Ofertas Compra",
            """SELECT entidade,morador,cpf,telefone,matricula,endereco,numero,tipo,
               num_siscor,data_siscor,opcao_compra,forma_pagamento,composicao_credito,
               form_correto,doc_identidade,vinculo_entidade,fatura_energia,fatura_agua,
               certidao_neg_1,certidao_neg_2,cadastro_cadunico,doc_herdeiro,
               num_ar_envio,data_entrega_form,dias_prorrog_form,status_resposta,
               valor_total,valor_pago,correcao_monetaria,outras_obs,
               data_limite_pgto,dias_prorrog_pgto,status_prazo_pgto,status_valor_pgto,
               data_email_escritura,valor_imovel,valor_com_desconto,
               data_pagamento,data_transferencia
               FROM ofertas_compra ORDER BY id""",
            ["Entidade","Morador","CPF","Telefone","Matricula","Endereco","Nr","Tipo",
             "SISCOR","Dt.SISCOR","Opcao","Forma Pgto","Composicao",
             "Form.OK","Doc.Ident.","Vinculo","Fat.Energia","Fat.Agua",
             "Cert.Neg.1","Cert.Neg.2","CadUnico","Doc.Herdeiro",
             "AR Envio","Dt.Entrega Form","Prorroga Form","Status Resposta",
             "Vlr.Total","Vlr.Pago","Correcao Mon.","Outras Obs",
             "Dt.Limite Pgto","Prorroga Pgto","Status Prazo","Status Valor",
             "Dt.Email Escritura","Vlr.Imovel","Vlr.c/Desconto",
             "Dt.Pagamento","Dt.Transferencia"])
