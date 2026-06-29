import tkinter as tk
from tkinter import ttk, messagebox
from database import get_conn
from widgets import tv_padrao, btn, FormDlg, AZUL, CINZA, BRANCO, VERMELHO, VERDE, LARANJA
from utils import (fmt_moeda, fmt_data, parse_moeda, parse_data,
                   OPCOES_COMPRA, FORMAS_PAGAMENTO, COMPOSICAO_CREDITO, STATUS_PARCELA_DOC)


class TabOfertas(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent, padding=12)
        self._build()
        self.carregar()

    def _build(self):
        topo = ttk.Frame(self)
        topo.pack(fill="x", pady=(0, 8))
        ttk.Label(topo, text="Ofertas de Compra 2026", style="Header.TLabel").pack(side="left")
        btn(topo, "+ Nova Oferta", self.novo, "Accent.TButton").pack(side="right", padx=3)
        btn(topo, "Editar", self.editar).pack(side="right", padx=3)
        btn(topo, "Excluir", self.excluir, "Danger.TButton").pack(side="right", padx=3)
        btn(topo, "Importar XLSX", self.importar).pack(side="right", padx=3)

        fil = ttk.Frame(self)
        fil.pack(fill="x", pady=(0, 6))
        ttk.Label(fil, text="Buscar:").pack(side="left")
        self.v_busca = tk.StringVar()
        self.v_busca.trace_add("write", lambda *_: self.carregar())
        ttk.Entry(fil, textvariable=self.v_busca, width=28).pack(side="left", padx=6)
        ttk.Label(fil, text="Status Pgto:").pack(side="left")
        self.v_spgto = tk.StringVar(value="TODOS")
        cb = ttk.Combobox(fil, textvariable=self.v_spgto,
                          values=["TODOS", "PGTO. FINALIZADO", "PENDENTE", "PRAZO EXPIRADO"],
                          width=18, state="readonly")
        cb.pack(side="left", padx=4)
        cb.bind("<<ComboboxSelected>>", lambda _: self.carregar())

        nb = ttk.Notebook(self)
        nb.pack(fill="both", expand=True)

        f1 = ttk.Frame(nb, padding=4)
        nb.add(f1, text="  Moradores / Formulario  ")
        cols1 = ["#", "Entidade", "Morador(a)", "CPF", "Telefone",
                 "Matricula", "Endereco", "Nr", "Tipo",
                 "SISCOR", "Data SISCOR", "Opcao", "Forma Pagto", "Composicao"]
        largs1 = [40, 160, 180, 110, 120, 75, 160, 45, 55, 100, 90, 110, 190, 180]
        self.tv1, frm1 = tv_padrao(f1, cols1, largs1, altura=14)
        frm1.pack(fill="both", expand=True)
        self.tv1.bind("<Double-1>", lambda _: self.editar())

        f2 = ttk.Frame(nb, padding=4)
        nb.add(f2, text="  Analise de Documentos  ")
        cols2 = ["Morador(a)", "Form.OK?", "Doc.Ident.", "Vinculos",
                 "Fat.Energia", "Fat.Agua", "Cert.Neg.1", "Cert.Neg.2",
                 "CadUnico", "Doc.Herdeiro"]
        largs2 = [180, 80, 90, 90, 90, 90, 100, 100, 80, 110]
        self.tv2, frm2 = tv_padrao(f2, cols2, largs2, altura=14)
        frm2.pack(fill="both", expand=True)

        f3 = ttk.Frame(nb, padding=4)
        nb.add(f3, text="  Prazos e Saldos  ")
        cols3 = ["Morador(a)", "AR Envio", "Dt.Entrega Form.", "Prorroga Form.",
                 "Status Resposta", "Valor Total", "Valor Pago",
                 "Correcao Mon.", "Dt.Limite Pgto", "Prorroga Pgto",
                 "Status Prazo", "Status Valor", "Dt.Email Escritura"]
        largs3 = [180, 120, 110, 80, 200, 110, 110, 110, 100, 80, 200, 160, 120]
        self.tv3, frm3 = tv_padrao(f3, cols3, largs3, altura=14)
        frm3.pack(fill="both", expand=True)

        f4 = ttk.Frame(nb, padding=4)
        nb.add(f4, text="  Informacoes Financeiras  ")
        cols4 = ["Morador(a)", "Matricula", "End.", "Nr",
                 "Valor Imovel", "Valor c/Desconto", "Valor Pago",
                 "Correcao Mon.", "Data Pagamento", "Data Transferencia"]
        largs4 = [180, 75, 160, 45, 110, 130, 110, 110, 100, 120]
        self.tv4, frm4 = tv_padrao(f4, cols4, largs4, altura=14)
        frm4.pack(fill="both", expand=True)

        self.v_total = tk.StringVar()
        ttk.Label(self, textvariable=self.v_total, foreground="#666").pack(anchor="w", pady=2)

        for tv in [self.tv1, self.tv2, self.tv3, self.tv4]:
            tv.bind("<<TreeviewSelect>>", self._sync_sel)
        self._tabs = [self.tv1, self.tv2, self.tv3, self.tv4]
        self._nb = nb

    def _sync_sel(self, event):
        src = event.widget
        sel = src.selection()
        if not sel: return
        iid = sel[0]
        for tv in self._tabs:
            if tv is not src:
                try: tv.selection_set(iid); tv.see(iid)
                except Exception: pass

    def carregar(self):
        for tv in self._tabs:
            for i in tv.get_children(): tv.delete(i)
        busca = f"%{self.v_busca.get()}%"
        spgto = self.v_spgto.get()
        with get_conn() as conn:
            q = """SELECT * FROM ofertas_compra WHERE
                   (morador LIKE ? OR entidade LIKE ? OR cpf LIKE ?
                    OR CAST(matricula AS TEXT) LIKE ?)"""
            p = [busca] * 4
            if spgto == "PGTO. FINALIZADO":
                q += " AND status_valor_pgto='PGTO. FINALIZADO'"
            elif spgto == "PENDENTE":
                q += " AND (status_valor_pgto IS NULL OR status_valor_pgto NOT IN ('PGTO. FINALIZADO'))"
            q += " ORDER BY id"
            rows = conn.execute(q, p).fetchall()
        self.v_total.set(f"{len(rows)} registro(s)")
        for i, r in enumerate(rows):
            tag = "pago" if r["status_valor_pgto"] == "PGTO. FINALIZADO" else ""
            iid = str(r["id"])
            self.tv1.insert("", "end", iid=iid, values=(
                i+1, r["entidade"] or "", r["morador"] or "",
                r["cpf"] or "", r["telefone"] or "",
                r["matricula"] or "", r["endereco"] or "", r["numero"] or "",
                r["tipo"] or "", r["num_siscor"] or "",
                fmt_data(r["data_siscor"]), r["opcao_compra"] or "",
                r["forma_pagamento"] or "", r["composicao_credito"] or ""
            ), tags=(tag,))
            self.tv2.insert("", "end", iid=iid, values=(
                r["morador"] or "", r["form_correto"] or "",
                r["doc_identidade"] or "", r["vinculo_entidade"] or "",
                r["fatura_energia"] or "", r["fatura_agua"] or "",
                r["certidao_neg_1"] or "", r["certidao_neg_2"] or "",
                r["cadastro_cadunico"] or "", r["doc_herdeiro"] or ""
            ), tags=(tag,))
            self.tv3.insert("", "end", iid=iid, values=(
                r["morador"] or "", r["num_ar_envio"] or "",
                fmt_data(r["data_entrega_form"]), r["dias_prorrog_form"] or "",
                r["status_resposta"] or "",
                fmt_moeda(r["valor_total"]) if r["valor_total"] else "",
                fmt_moeda(r["valor_pago"]) if r["valor_pago"] else "",
                fmt_moeda(r["correcao_monetaria"]) if r["correcao_monetaria"] else "",
                fmt_data(r["data_limite_pgto"]), r["dias_prorrog_pgto"] or "",
                r["status_prazo_pgto"] or "", r["status_valor_pgto"] or "",
                fmt_data(r["data_email_escritura"])
            ), tags=(tag,))
            self.tv4.insert("", "end", iid=iid, values=(
                r["morador"] or "", r["matricula"] or "",
                r["endereco"] or "", r["numero"] or "",
                fmt_moeda(r["valor_imovel"]) if r["valor_imovel"] else "",
                fmt_moeda(r["valor_com_desconto"]) if r["valor_com_desconto"] else "",
                fmt_moeda(r["valor_pago"]) if r["valor_pago"] else "",
                fmt_moeda(r["correcao_monetaria"]) if r["correcao_monetaria"] else "",
                fmt_data(r["data_pagamento"]), fmt_data(r["data_transferencia"])
            ), tags=(tag,))

    def _sel(self):
        for tv in self._tabs:
            sel = tv.selection()
            if sel: return int(sel[0])
        messagebox.showwarning("Atencao", "Selecione um registro.")
        return None

    def novo(self):
        d = OfertaDlg(self); self.wait_window(d); self.carregar()

    def editar(self):
        oid = self._sel()
        if oid is None: return
        with get_conn() as conn:
            row = conn.execute("SELECT * FROM ofertas_compra WHERE id=?", (oid,)).fetchone()
        d = OfertaDlg(self, row); self.wait_window(d); self.carregar()

    def excluir(self):
        oid = self._sel()
        if oid is None: return
        if messagebox.askyesno("Confirmar", "Excluir esta oferta de compra?"):
            with get_conn() as conn:
                conn.execute("DELETE FROM ofertas_compra WHERE id=?", (oid,))
            self.carregar()

    def importar(self):
        from tkinter.filedialog import askopenfilename
        path = askopenfilename(title="Selecionar planilha de ofertas",
                               filetypes=[("Excel", "*.xlsx *.xlsm"), ("Todos", "*.*")])
        if not path: return
        ImportarOfertasDlg(self, path)
        self.carregar()


class OfertaDlg(FormDlg):
    def __init__(self, parent, row=None):
        super().__init__(parent, "Editar Oferta" if row else "Nova Oferta de Compra", w=760, h=700)
        self.row = row; self._build()
        if row: self._fill()

    def _build(self):
        nb = ttk.Notebook(self.corpo)
        nb.pack(fill="both", expand=True)
        t1 = ttk.Frame(nb, padding=8); nb.add(t1, text="Dados do Morador")
        for i in range(4): t1.columnconfigure(i, weight=1)
        self.v_ent, _ = self.campo(t1, "Entidade", 0, 0, cs=2, w=40)
        self.v_mor, _ = self.campo(t1, "Morador(a) *", 0, 2, cs=2, w=40)
        self.v_cpf, _ = self.campo(t1, "CPF", 1, 0)
        self.v_tel, _ = self.campo(t1, "Telefone", 1, 1)
        self.v_email, _ = self.campo(t1, "E-mail", 1, 2, cs=2, w=40)
        self.v_mat, _ = self.campo(t1, "Matricula", 2, 0)
        self.v_end, _ = self.campo(t1, "Endereco", 2, 1, cs=2, w=40)
        self.v_num, _ = self.campo(t1, "Nr", 2, 3)
        self.v_tipo, _ = self.campo(t1, "Tipo", 3, 0)
        self.v_siscor, _ = self.campo(t1, "Nr SISCOR Resposta", 3, 1)
        self.v_dsiscor, _ = self.campo(t1, "Data SISCOR (dd/mm/aaaa)", 3, 2)
        self.v_opcao, _ = self.combo(t1, "Opcao de Compra", OPCOES_COMPRA, 4, 0)
        self.v_forma, _ = self.combo(t1, "Forma de Pagamento", FORMAS_PAGAMENTO, 4, 1, cs=2)
        self.v_comp, _ = self.combo(t1, "Composicao do Credito", COMPOSICAO_CREDITO, 4, 3)

        t2 = ttk.Frame(nb, padding=8); nb.add(t2, text="Analise de Documentos")
        for i in range(4): t2.columnconfigure(i, weight=1)
        op = STATUS_PARCELA_DOC
        self.v_form_ok, _ = self.combo(t2, "Formulario Correto?", ["SIM", "NAO"], 0, 0)
        self.v_doc_id, _ = self.combo(t2, "Doc. Identidade", op, 0, 1)
        self.v_vinculo, _ = self.combo(t2, "Vinculo com Entidade", op, 0, 2)
        self.v_energ, _ = self.combo(t2, "Fat. Energia (3 meses)", op, 1, 0)
        self.v_agua, _ = self.combo(t2, "Fat. Agua/Esgoto (3 meses)", op, 1, 1)
        self.v_cert1, _ = self.combo(t2, "Cert. Neg. 1 Oficio", op, 1, 2)
        self.v_cert2, _ = self.combo(t2, "Cert. Neg. 2 Oficio", op, 1, 3)
        self.v_cadunico, _ = self.combo(t2, "CadUnico?", ["SIM", "NAO"], 2, 0)
        self.v_herdeiro, _ = self.combo(t2, "Doc. Identidade Herdeiro(a)", op, 2, 1)

        t3 = ttk.Frame(nb, padding=8); nb.add(t3, text="Prazos e Saldos")
        for i in range(4): t3.columnconfigure(i, weight=1)
        self.v_ar, _ = self.campo(t3, "Nr AR Envio Formulario", 0, 0)
        self.v_dt_form, _ = self.campo(t3, "Data Entrega Form. (dd/mm/aaaa)", 0, 1)
        self.v_prorrform, _ = self.campo(t3, "Dias Prorrogacao (form)", 0, 2)
        self.v_st_resp, _ = self.campo(t3, "Status Resposta Formulario", 1, 0, cs=4, w=60)
        self.v_vltotal, _ = self.campo(t3, "Valor Total a Pagar (R$)", 2, 0)
        self.v_vlpago, _ = self.campo(t3, "Valor Pago (R$)", 2, 1)
        self.v_corr, _ = self.campo(t3, "Correcao Monetaria (R$)", 2, 2)
        self.v_outras, _ = self.campo(t3, "Outras Obs.", 2, 3)
        self.v_dt_pgto, _ = self.campo(t3, "Data Limite Pagamento (dd/mm/aaaa)", 3, 0)
        self.v_prorrpgto, _ = self.campo(t3, "Dias Prorrogacao (pgto)", 3, 1)
        self.v_st_prazo, _ = self.campo(t3, "Status Prazo Pgto", 4, 0, cs=2, w=50)
        self.v_st_valor, _ = self.campo(t3, "Status Valor Pgto", 4, 2, cs=2, w=50)
        self.v_dt_email, _ = self.campo(t3, "Data Envio E-mail Escritura (dd/mm/aaaa)", 5, 0, cs=2)

        t4 = ttk.Frame(nb, padding=8); nb.add(t4, text="Informacoes Financeiras")
        for i in range(4): t4.columnconfigure(i, weight=1)
        self.v_vlimovel, _ = self.campo(t4, "Valor do Imovel (R$)", 0, 0)
        self.v_vldesc, _ = self.campo(t4, "Valor com Desconto (R$)", 0, 1)
        self.v_vlpago2, _ = self.campo(t4, "Valor Pago (R$)", 0, 2)
        self.v_corrfin, _ = self.campo(t4, "Correcao Monetaria (R$)", 0, 3)
        self.v_dt_pag, _ = self.campo(t4, "Data Pagamento (dd/mm/aaaa)", 1, 0)
        self.v_dt_transf, _ = self.campo(t4, "Data Transferencia (dd/mm/aaaa)", 1, 1)
        self.t_obs = self.texto(t4, "Observacoes Gerais", 2, 0, cs=4, h=4)

        btn(self.rodape, "Salvar", self._salvar, "Accent.TButton").pack(side="right", padx=8, pady=6)
        btn(self.rodape, "Cancelar", self.destroy).pack(side="right", pady=6)

    def _fill(self):
        r = self.row
        self.v_ent.set(r["entidade"] or ""); self.v_mor.set(r["morador"] or "")
        self.v_cpf.set(r["cpf"] or ""); self.v_tel.set(r["telefone"] or "")
        self.v_email.set(r["email"] or ""); self.v_mat.set(r["matricula"] or "")
        self.v_end.set(r["endereco"] or ""); self.v_num.set(r["numero"] or "")
        self.v_tipo.set(r["tipo"] or ""); self.v_siscor.set(r["num_siscor"] or "")
        self.v_dsiscor.set(fmt_data(r["data_siscor"]))
        self.v_opcao.set(r["opcao_compra"] or OPCOES_COMPRA[0])
        self.v_forma.set(r["forma_pagamento"] or FORMAS_PAGAMENTO[0])
        self.v_comp.set(r["composicao_credito"] or COMPOSICAO_CREDITO[0])
        self.v_form_ok.set(r["form_correto"] or "NAO")
        self.v_doc_id.set(r["doc_identidade"] or "NAO SE APLICA")
        self.v_vinculo.set(r["vinculo_entidade"] or "NAO SE APLICA")
        self.v_energ.set(r["fatura_energia"] or "NAO"); self.v_agua.set(r["fatura_agua"] or "NAO")
        self.v_cert1.set(r["certidao_neg_1"] or "NAO"); self.v_cert2.set(r["certidao_neg_2"] or "NAO")
        self.v_cadunico.set(r["cadastro_cadunico"] or "NAO")
        self.v_herdeiro.set(r["doc_herdeiro"] or "NAO SE APLICA")
        self.v_ar.set(r["num_ar_envio"] or ""); self.v_dt_form.set(fmt_data(r["data_entrega_form"]))
        self.v_prorrform.set(r["dias_prorrog_form"] or ""); self.v_st_resp.set(r["status_resposta"] or "")
        self.v_vltotal.set(r["valor_total"] or ""); self.v_vlpago.set(r["valor_pago"] or "")
        self.v_corr.set(r["correcao_monetaria"] or ""); self.v_outras.set(r["outras_obs"] or "")
        self.v_dt_pgto.set(fmt_data(r["data_limite_pgto"])); self.v_prorrpgto.set(r["dias_prorrog_pgto"] or "")
        self.v_st_prazo.set(r["status_prazo_pgto"] or ""); self.v_st_valor.set(r["status_valor_pgto"] or "")
        self.v_dt_email.set(fmt_data(r["data_email_escritura"]))
        self.v_vlimovel.set(r["valor_imovel"] or ""); self.v_vldesc.set(r["valor_com_desconto"] or "")
        self.v_vlpago2.set(r["valor_pago"] or ""); self.v_corrfin.set(r["correcao_monetaria"] or "")
        self.v_dt_pag.set(fmt_data(r["data_pagamento"])); self.v_dt_transf.set(fmt_data(r["data_transferencia"]))
        self.t_obs.insert("1.0", r["observacoes"] or "")

    def _salvar(self):
        mor = self.v_mor.get().strip()
        if not mor: messagebox.showerror("Erro", "Nome do morador e obrigatorio."); return
        d = dict(
            entidade=self.v_ent.get().strip() or None, morador=mor,
            cpf=self.v_cpf.get().strip() or None, telefone=self.v_tel.get().strip() or None,
            email=self.v_email.get().strip() or None, matricula=self.v_mat.get().strip() or None,
            endereco=self.v_end.get().strip() or None, numero=self.v_num.get().strip() or None,
            tipo=self.v_tipo.get().strip() or None, num_siscor=self.v_siscor.get().strip() or None,
            data_siscor=parse_data(self.v_dsiscor.get()),
            opcao_compra=self.v_opcao.get(), forma_pagamento=self.v_forma.get(),
            composicao_credito=self.v_comp.get(), form_correto=self.v_form_ok.get(),
            doc_identidade=self.v_doc_id.get(), vinculo_entidade=self.v_vinculo.get(),
            fatura_energia=self.v_energ.get(), fatura_agua=self.v_agua.get(),
            certidao_neg_1=self.v_cert1.get(), certidao_neg_2=self.v_cert2.get(),
            cadastro_cadunico=self.v_cadunico.get(), doc_herdeiro=self.v_herdeiro.get(),
            num_ar_envio=self.v_ar.get().strip() or None,
            data_entrega_form=parse_data(self.v_dt_form.get()),
            dias_prorrog_form=self.v_prorrform.get().strip() or 0,
            status_resposta=self.v_st_resp.get().strip() or None,
            valor_total=parse_moeda(self.v_vltotal.get()) or None,
            valor_pago=parse_moeda(self.v_vlpago.get()) or None,
            correcao_monetaria=parse_moeda(self.v_corr.get()) or None,
            outras_obs=self.v_outras.get().strip() or None,
            data_limite_pgto=parse_data(self.v_dt_pgto.get()),
            dias_prorrog_pgto=self.v_prorrpgto.get().strip() or 0,
            status_prazo_pgto=self.v_st_prazo.get().strip() or None,
            status_valor_pgto=self.v_st_valor.get().strip() or None,
            data_email_escritura=parse_data(self.v_dt_email.get()),
            valor_imovel=parse_moeda(self.v_vlimovel.get()) or None,
            valor_com_desconto=parse_moeda(self.v_vldesc.get()) or None,
            data_pagamento=parse_data(self.v_dt_pag.get()),
            data_transferencia=parse_data(self.v_dt_transf.get()),
            observacoes=self.t_obs.get("1.0", "end-1c").strip() or None
        )
        with get_conn() as conn:
            if self.row:
                cols = ", ".join(f"{k}=:{k}" for k in d)
                conn.execute(f"UPDATE ofertas_compra SET {cols} WHERE id=?", {**d, "?": self.row["id"]})
            else:
                ks = ", ".join(d.keys())
                vs = ", ".join(f":{k}" for k in d.keys())
                conn.execute(f"INSERT INTO ofertas_compra ({ks}) VALUES ({vs})", d)
        self.destroy()


class ImportarOfertasDlg(tk.Toplevel):
    def __init__(self, parent, path):
        super().__init__(parent)
        self.title("Importar Ofertas de Compra - XLSX")
        self.geometry("700x500")
        self.grab_set()
        self.path = path
        self._build()
        self.after(100, self._carregar_abas)

    def _build(self):
        ttk.Label(self, text=f"Arquivo: {self.path.split('/')[-1]}",
                  font=("Segoe UI", 10)).pack(anchor="w", padx=12, pady=8)
        f = ttk.Frame(self); f.pack(fill="x", padx=12)
        ttk.Label(f, text="Aba:").pack(side="left")
        self.v_aba = tk.StringVar()
        self.cb_aba = ttk.Combobox(f, textvariable=self.v_aba, state="readonly", width=35)
        self.cb_aba.pack(side="left", padx=6)
        self.log = tk.Text(self, height=18, font=("Consolas", 9))
        self.log.pack(fill="both", expand=True, padx=12, pady=4)
        rodape = ttk.Frame(self); rodape.pack(fill="x", padx=12, pady=6)
        btn(rodape, "Importar Dados Moradores", self._import_moradores, "Accent.TButton").pack(side="left", padx=4)
        btn(rodape, "Importar Prazos/Saldos", self._import_prazos).pack(side="left", padx=4)
        btn(rodape, "Importar Financeiro", self._import_financeiro).pack(side="left", padx=4)
        btn(rodape, "Fechar", self.destroy).pack(side="right")

    def _log(self, t): self.log.insert("end", t + "\n"); self.log.see("end")

    def _carregar_abas(self):
        import openpyxl
        wb = openpyxl.load_workbook(self.path, read_only=True, data_only=True, keep_vba=False)
        self.cb_aba["values"] = wb.sheetnames
        if "Dados Moradores" in wb.sheetnames: self.cb_aba.set("Dados Moradores")
        elif wb.sheetnames: self.cb_aba.set(wb.sheetnames[0])
        wb.close()

    def _get_rows(self):
        import openpyxl
        aba = self.v_aba.get()
        wb = openpyxl.load_workbook(self.path, read_only=True, data_only=True, keep_vba=False)
        ws = wb[aba]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        return rows

    def _ci(self, header, *nomes):
        for nm in nomes:
            for j, v in enumerate(header):
                if nm in v: return j
        return None

    def _import_moradores(self):
        rows = self._get_rows()
        header = None; start = 0
        for i, row in enumerate(rows):
            vals = [str(v).upper().strip() if v else "" for v in row]
            if "MORADOR(A)" in vals or "MORADOR" in vals:
                header = vals; start = i + 1; break
        if not header: self._log("Cabecalho nao encontrado."); return
        ci = self._ci
        i_ent = ci(header, "ENTIDADE"); i_mor = ci(header, "MORADOR")
        i_cpf = ci(header, "CPF"); i_tel = ci(header, "TELEFONE")
        i_email = ci(header, "E-MAIL", "EMAIL"); i_mat = ci(header, "MATRICULA")
        i_end = ci(header, "ENDERECO"); i_num = ci(header, "NR", "NUMERO")
        i_tipo = ci(header, "TIPO")
        ok = skip = 0
        with get_conn() as conn:
            for row in rows[start:]:
                if all(v is None for v in row): continue
                def g(i): return row[i] if i is not None and i < len(row) else None
                mor = str(g(i_mor) or "").strip()
                if not mor or mor == "None": skip += 1; continue
                try:
                    conn.execute("""INSERT OR IGNORE INTO ofertas_compra
                        (entidade,morador,cpf,telefone,email,matricula,endereco,numero,tipo)
                        VALUES (?,?,?,?,?,?,?,?,?)""",
                        (str(g(i_ent) or "").strip() or None, mor,
                         str(g(i_cpf) or "").strip() or None, str(g(i_tel) or "").strip() or None,
                         str(g(i_email) or "").strip() or None, g(i_mat),
                         str(g(i_end) or "").strip() or None, str(g(i_num) or "").strip() or None,
                         str(g(i_tipo) or "").strip() or None))
                    ok += 1
                except Exception as e:
                    self._log(f"  Erro {mor}: {e}"); skip += 1
        self._log(f"Moradores: {ok} inseridos, {skip} ignorados.")

    def _import_prazos(self):
        rows = self._get_rows()
        header = None; start = 0
        for i, row in enumerate(rows):
            vals = [str(v).upper().strip() if v else "" for v in row]
            if "STATUS RESPOSTA" in " ".join(vals) or "VALOR TOTAL" in " ".join(vals):
                header = vals; start = i + 1; break
        if not header: self._log("Aba de Prazos: cabecalho nao encontrado."); return
        ci = self._ci
        i_mor = ci(header, "MORADOR"); i_mat = ci(header, "MATRICULA")
        i_ar = ci(header, "AR"); i_dtform = ci(header, "DATA ENTREGA")
        i_prorrform = ci(header, "DIAS"); i_stresp = ci(header, "STATUS RESPOSTA")
        i_vltot = ci(header, "VALOR TOTAL"); i_vlpago = ci(header, "VALOR PAGO")
        i_corr = ci(header, "CORRECAO"); i_outras = ci(header, "OUTRAS")
        i_dtpgto = ci(header, "DATA FINAL", "DATA LIMITE"); i_prorrpgto = ci(header, "DIAS")
        i_stprazo = ci(header, "STATUS PRAZO"); i_stvalor = ci(header, "STATUS VALOR")
        i_dtemail = ci(header, "DATA ENVIO")
        ok = skip = 0
        with get_conn() as conn:
            for row in rows[start:]:
                if all(v is None for v in row): continue
                def g(i): return row[i] if i is not None and i < len(row) else None
                mor = str(g(i_mor) or "").strip(); mat = g(i_mat)
                if not mor or mor == "None": skip += 1; continue
                dtform = g(i_dtform)
                if hasattr(dtform, "strftime"): dtform = dtform.strftime("%Y-%m-%d")
                dtpgto = g(i_dtpgto)
                if hasattr(dtpgto, "strftime"): dtpgto = dtpgto.strftime("%Y-%m-%d")
                dtemail = g(i_dtemail)
                if hasattr(dtemail, "strftime"): dtemail = dtemail.strftime("%Y-%m-%d")
                existing = conn.execute("SELECT id FROM ofertas_compra WHERE morador=? OR (matricula=? AND matricula IS NOT NULL) LIMIT 1", (mor, mat)).fetchone()
                if existing:
                    conn.execute("""UPDATE ofertas_compra SET num_ar_envio=?,data_entrega_form=?,
                        dias_prorrog_form=?,status_resposta=?,valor_total=?,valor_pago=?,
                        correcao_monetaria=?,outras_obs=?,data_limite_pgto=?,dias_prorrog_pgto=?,
                        status_prazo_pgto=?,status_valor_pgto=?,data_email_escritura=? WHERE id=?""",
                        (str(g(i_ar) or "").strip() or None, dtform, g(i_prorrform),
                         str(g(i_stresp) or "").strip() or None, g(i_vltot), g(i_vlpago),
                         g(i_corr), str(g(i_outras) or "").strip() or None, dtpgto,
                         g(i_prorrpgto), str(g(i_stprazo) or "").strip() or None,
                         str(g(i_stvalor) or "").strip() or None, dtemail, existing["id"]))
                    ok += 1
                else: skip += 1
        self._log(f"Prazos: {ok} atualizados, {skip} nao encontrados.")

    def _import_financeiro(self):
        rows = self._get_rows()
        header = None; start = 0
        for i, row in enumerate(rows):
            vals = [str(v).upper().strip() if v else "" for v in row]
            if "VALOR DO" in " ".join(vals) or "VALOR COM" in " ".join(vals):
                header = vals; start = i + 1; break
        if not header: self._log("Aba Financeira: cabecalho nao encontrado."); return
        ci = self._ci
        i_mor = ci(header, "MORADOR"); i_mat = ci(header, "MATRICULA")
        i_vlimovel = ci(header, "VALOR DO", "VALOR IMOVEL")
        i_vldesc = ci(header, "VALOR COM"); i_vlpago = ci(header, "VALOR PAGO")
        i_corr = ci(header, "CORRECAO"); i_dtpag = ci(header, "DATA PAG")
        i_dttransf = ci(header, "DATA DA", "DATA TRANSFERENCIA")
        ok = skip = 0
        with get_conn() as conn:
            for row in rows[start:]:
                if all(v is None for v in row): continue
                def g(i): return row[i] if i is not None and i < len(row) else None
                mor = str(g(i_mor) or "").strip(); mat = g(i_mat)
                if not mor or mor == "None": skip += 1; continue
                dtpag = g(i_dtpag)
                if hasattr(dtpag, "strftime"): dtpag = dtpag.strftime("%Y-%m-%d")
                dttransf = g(i_dttransf)
                if hasattr(dttransf, "strftime"): dttransf = dttransf.strftime("%Y-%m-%d")
                existing = conn.execute("SELECT id FROM ofertas_compra WHERE morador=? OR (matricula=? AND matricula IS NOT NULL) LIMIT 1", (mor, mat)).fetchone()
                if existing:
                    conn.execute("UPDATE ofertas_compra SET valor_imovel=?,valor_com_desconto=?,valor_pago=?,correcao_monetaria=?,data_pagamento=?,data_transferencia=? WHERE id=?",
                        (g(i_vlimovel), g(i_vldesc), g(i_vlpago), g(i_corr), dtpag, dttransf, existing["id"]))
                    ok += 1
                else: skip += 1
        self._log(f"Financeiro: {ok} atualizados, {skip} nao encontrados.")
