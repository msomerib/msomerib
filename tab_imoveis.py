import tkinter as tk
from tkinter import ttk, messagebox
from database import get_conn
from widgets import tv_padrao, btn, FormDlg, AZUL, CINZA, VERMELHO, VERDE, LARANJA
from utils import fmt_moeda, fmt_data, parse_moeda, parse_data, STATUS_IMOVEL


class TabImoveis(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent, padding=12)
        self._build()
        self.carregar()

    def _build(self):
        topo = ttk.Frame(self)
        topo.pack(fill="x", pady=(0, 8))
        ttk.Label(topo, text="Cadastro de Imoveis", style="Header.TLabel").pack(side="left")
        btn(topo, "+ Novo", self.novo, "Accent.TButton").pack(side="right", padx=3)
        btn(topo, "Editar", self.editar).pack(side="right", padx=3)
        btn(topo, "Excluir", self.excluir, "Danger.TButton").pack(side="right", padx=3)
        btn(topo, "Importar XLSX", self.importar).pack(side="right", padx=3)

        fil = ttk.Frame(self)
        fil.pack(fill="x", pady=(0, 6))
        ttk.Label(fil, text="Buscar:").pack(side="left")
        self.v_busca = tk.StringVar()
        self.v_busca.trace_add("write", lambda *_: self.carregar())
        ttk.Entry(fil, textvariable=self.v_busca, width=30).pack(side="left", padx=6)
        ttk.Label(fil, text="Status:").pack(side="left")
        self.v_status = tk.StringVar(value="TODOS")
        cb = ttk.Combobox(fil, textvariable=self.v_status,
                          values=["TODOS"] + STATUS_IMOVEL, width=14, state="readonly")
        cb.pack(side="left", padx=6)
        cb.bind("<<ComboboxSelected>>", lambda _: self.carregar())
        ttk.Label(fil, text="Contrato:").pack(side="left")
        self.v_tipo_ct = tk.StringVar(value="TODOS")
        cb2 = ttk.Combobox(fil, textvariable=self.v_tipo_ct,
                           values=["TODOS", "ONEROSO", "COMODATO", "CESSAO GRATUITA"],
                           width=14, state="readonly")
        cb2.pack(side="left", padx=6)
        cb2.bind("<<ComboboxSelected>>", lambda _: self.carregar())

        cols = ["Codigo", "Entidade / Morador", "Endereco", "Nr", "Q", "L",
                "Area m2", "Matricula", "Contrato", "Tipo CT", "Vigencia",
                "Sit. Contrato", "Tx.Ocup.", "Preco Min.", "Status"]
        largs = [80, 160, 160, 50, 50, 50, 75, 85, 120, 90, 90, 90, 85, 110, 90]
        self.tv, frm = tv_padrao(self, cols, largs, altura=20)
        frm.pack(fill="both", expand=True)
        self.tv.bind("<Double-1>", lambda _: self.editar())

        self.v_total = tk.StringVar(value="")
        ttk.Label(self, textvariable=self.v_total, foreground="#666").pack(anchor="w", pady=2)

    def carregar(self):
        for i in self.tv.get_children():
            self.tv.delete(i)
        busca = f"%{self.v_busca.get()}%"
        status = self.v_status.get()
        tipo_ct = self.v_tipo_ct.get()
        with get_conn() as conn:
            q = """SELECT * FROM imoveis WHERE
                   (codigo LIKE ? OR entidade LIKE ? OR morador LIKE ?
                    OR endereco LIKE ? OR CAST(matricula AS TEXT) LIKE ?)"""
            p = [busca] * 5
            if status != "TODOS":
                q += " AND status=?"; p.append(status)
            if tipo_ct != "TODOS":
                q += " AND tipo_contrato=?"; p.append(tipo_ct)
            q += " ORDER BY codigo"
            rows = conn.execute(q, p).fetchall()
        self.v_total.set(f"{len(rows)} registro(s)")
        for r in rows:
            sit = r["sit_contrato"] or ""
            tag = "vencido" if sit == "VENCIDO" else ("vigente" if sit == "VIGENTE" else "")
            self.tv.insert("", "end", iid=r["id"], values=(
                r["codigo"], r["entidade"] or r["morador"] or "",
                r["endereco"], r["numero"] or "", r["quadra"] or "", r["lote"] or "",
                r["area_m2"] or "", r["matricula"] or "",
                r["num_contrato"] or "", r["tipo_contrato"] or "",
                fmt_data(r["vigencia"]), sit,
                fmt_moeda(r["taxa_ocupacao"]) if r["taxa_ocupacao"] else "",
                fmt_moeda(r["preco_minimo"]) if r["preco_minimo"] else "",
                r["status"]
            ), tags=(tag,))

    def _sel(self):
        sel = self.tv.selection()
        if not sel:
            messagebox.showwarning("Atencao", "Selecione um imovel.")
            return None
        return int(sel[0])

    def novo(self):
        d = ImovelDlg(self)
        self.wait_window(d)
        self.carregar()

    def editar(self):
        iid = self._sel()
        if iid is None:
            return
        with get_conn() as conn:
            row = conn.execute("SELECT * FROM imoveis WHERE id=?", (iid,)).fetchone()
        d = ImovelDlg(self, row)
        self.wait_window(d)
        self.carregar()

    def excluir(self):
        iid = self._sel()
        if iid is None:
            return
        if messagebox.askyesno("Confirmar", "Excluir este imovel permanentemente?"):
            with get_conn() as conn:
                conn.execute("DELETE FROM imoveis WHERE id=?", (iid,))
            self.carregar()

    def importar(self):
        from tkinter.filedialog import askopenfilename
        path = askopenfilename(title="Selecionar planilha Excel",
                               filetypes=[("Excel", "*.xlsx *.xlsm"), ("Todos", "*.*")])
        if not path:
            return
        ImportarImovelDlg(self, path)
        self.carregar()


class ImovelDlg(FormDlg):
    def __init__(self, parent, row=None):
        super().__init__(parent, "Editar Imovel" if row else "Novo Imovel", w=700, h=620)
        self.row = row
        self._build()
        if row:
            self._fill()

    def _build(self):
        c = self.corpo
        for i in range(4):
            c.columnconfigure(i, weight=1)
        self.v_cod, _ = self.campo(c, "Codigo (ex: A-1129) *", 0, 0)
        self.v_entidade, _ = self.campo(c, "Entidade", 0, 1, cs=2, w=40)
        self.v_morador, _ = self.campo(c, "Morador(a)", 0, 3)
        self.v_end, _ = self.campo(c, "Endereco *", 1, 0, cs=3, w=50)
        self.v_num, _ = self.campo(c, "Nr", 1, 3, w=10)
        self.v_qd, _ = self.campo(c, "Quadra", 2, 0, w=10)
        self.v_lote, _ = self.campo(c, "Lote", 2, 1, w=10)
        self.v_tipo, _ = self.campo(c, "Tipo Imovel", 2, 2, w=10)
        self.v_local, _ = self.campo(c, "Local/Area", 2, 3, w=10)
        self.v_area, _ = self.campo(c, "Area m2", 3, 0, w=14)
        self.v_constr, _ = self.campo(c, "Constr. m2", 3, 1, w=14)
        self.v_mat, _ = self.campo(c, "Matricula", 3, 2, w=14)
        self.v_status, _ = self.combo(c, "Status", STATUS_IMOVEL, 3, 3, w=14)
        self.v_nct, _ = self.campo(c, "Nr Contrato", 4, 0, w=22)
        self.v_tct, _ = self.combo(c, "Tipo Contrato", ["", "ONEROSO", "COMODATO", "CESSAO GRATUITA"], 4, 1, w=16)
        self.v_vig, _ = self.campo(c, "Vigencia (dd/mm/aaaa)", 4, 2, w=16)
        self.v_sit, _ = self.combo(c, "Sit. Contrato", ["", "VIGENTE", "VENCIDO", "RESCINDIDO"], 4, 3, w=14)
        self.v_txoc, _ = self.campo(c, "Taxa Ocupacao (R$)", 5, 0, w=18)
        self.v_pmin, _ = self.campo(c, "Preco Min. Alienacao (R$)", 5, 1, w=18)
        self.v_p25, _ = self.campo(c, "Preco 25% (R$)", 5, 2, w=18)
        self.t_obs = self.texto(c, "Observacoes", 6, 0, cs=4, h=3)
        btn(self.rodape, "Salvar", self._salvar, "Accent.TButton").pack(side="right", padx=8, pady=6)
        btn(self.rodape, "Cancelar", self.destroy).pack(side="right", pady=6)

    def _fill(self):
        r = self.row
        self.v_cod.set(r["codigo"])
        self.v_entidade.set(r["entidade"] or "")
        self.v_morador.set(r["morador"] or "")
        self.v_end.set(r["endereco"])
        self.v_num.set(r["numero"] or "")
        self.v_qd.set(r["quadra"] or "")
        self.v_lote.set(r["lote"] or "")
        self.v_tipo.set(r["tipo"] or "")
        self.v_local.set(r["local_area"] or "")
        self.v_area.set(r["area_m2"] or "")
        self.v_constr.set(r["construcao_m2"] or "")
        self.v_mat.set(r["matricula"] or "")
        self.v_status.set(r["status"] or "OCUPADO")
        self.v_nct.set(r["num_contrato"] or "")
        self.v_tct.set(r["tipo_contrato"] or "")
        self.v_vig.set(fmt_data(r["vigencia"]))
        self.v_sit.set(r["sit_contrato"] or "")
        self.v_txoc.set(r["taxa_ocupacao"] or "")
        self.v_pmin.set(r["preco_minimo"] or "")
        self.v_p25.set(r["preco_25pct"] or "")
        self.t_obs.insert("1.0", r["observacoes"] or "")

    def _salvar(self):
        cod = self.v_cod.get().strip()
        end = self.v_end.get().strip()
        if not cod or not end:
            messagebox.showerror("Erro", "Codigo e Endereco sao obrigatorios.")
            return
        d = dict(
            codigo=cod, entidade=self.v_entidade.get().strip(),
            morador=self.v_morador.get().strip(),
            endereco=end, numero=self.v_num.get().strip(),
            quadra=self.v_qd.get().strip() or None,
            lote=self.v_lote.get().strip() or None,
            tipo=self.v_tipo.get().strip(),
            local_area=self.v_local.get().strip(),
            area_m2=self.v_area.get().strip() or None,
            construcao_m2=self.v_constr.get().strip() or None,
            matricula=self.v_mat.get().strip() or None,
            status=self.v_status.get(),
            num_contrato=self.v_nct.get().strip() or None,
            tipo_contrato=self.v_tct.get() or None,
            vigencia=parse_data(self.v_vig.get()),
            sit_contrato=self.v_sit.get() or None,
            taxa_ocupacao=parse_moeda(self.v_txoc.get()) or None,
            preco_minimo=parse_moeda(self.v_pmin.get()) or None,
            preco_25pct=parse_moeda(self.v_p25.get()) or None,
            observacoes=self.t_obs.get("1.0", "end-1c").strip()
        )
        with get_conn() as conn:
            if self.row:
                conn.execute("""UPDATE imoveis SET codigo=:codigo, entidade=:entidade,
                    morador=:morador, endereco=:endereco, numero=:numero, quadra=:quadra,
                    lote=:lote, tipo=:tipo, local_area=:local_area, area_m2=:area_m2,
                    construcao_m2=:construcao_m2, matricula=:matricula, status=:status,
                    num_contrato=:num_contrato, tipo_contrato=:tipo_contrato,
                    vigencia=:vigencia, sit_contrato=:sit_contrato,
                    taxa_ocupacao=:taxa_ocupacao, preco_minimo=:preco_minimo,
                    preco_25pct=:preco_25pct, observacoes=:observacoes
                    WHERE id=?""", {**d, "?": self.row["id"]})
            else:
                conn.execute("""INSERT INTO imoveis (codigo,entidade,morador,endereco,numero,
                    quadra,lote,tipo,local_area,area_m2,construcao_m2,matricula,status,
                    num_contrato,tipo_contrato,vigencia,sit_contrato,taxa_ocupacao,
                    preco_minimo,preco_25pct,observacoes) VALUES
                    (:codigo,:entidade,:morador,:endereco,:numero,:quadra,:lote,:tipo,
                    :local_area,:area_m2,:construcao_m2,:matricula,:status,:num_contrato,
                    :tipo_contrato,:vigencia,:sit_contrato,:taxa_ocupacao,:preco_minimo,
                    :preco_25pct,:observacoes)""", d)
        self.destroy()


class ImportarImovelDlg(tk.Toplevel):
    def __init__(self, parent, path):
        super().__init__(parent)
        self.title("Importar Imoveis - XLSX")
        self.geometry("700x500")
        self.grab_set()
        self.path = path
        self._build()
        self.after(100, self._carregar_preview)

    def _build(self):
        ttk.Label(self, text=f"Arquivo: {self.path.split('/')[-1]}",
                  font=("Segoe UI", 10)).pack(anchor="w", padx=12, pady=8)
        ttk.Label(self, text="Aba:").pack(anchor="w", padx=12)
        self.v_aba = tk.StringVar()
        self.cb_aba = ttk.Combobox(self, textvariable=self.v_aba, state="readonly", width=30)
        self.cb_aba.pack(anchor="w", padx=12, pady=4)
        self.cb_aba.bind("<<ComboboxSelected>>", lambda _: self._preview_aba())
        self.log = tk.Text(self, height=18, font=("Consolas", 9))
        self.log.pack(fill="both", expand=True, padx=12, pady=4)
        rodape = ttk.Frame(self)
        rodape.pack(fill="x", padx=12, pady=6)
        btn(rodape, "Importar Selecionada", self._importar, "Accent.TButton").pack(side="right", padx=4)
        btn(rodape, "Fechar", self.destroy).pack(side="right")
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=False)
        self.cb_aba["values"] = wb.sheetnames
        if wb.sheetnames:
            self.cb_aba.set(wb.sheetnames[0])
        wb.close()

    def _log(self, txt):
        self.log.insert("end", txt + "\n")
        self.log.see("end")

    def _carregar_preview(self):
        self._preview_aba()

    def _preview_aba(self):
        self.log.delete("1.0", "end")
        import openpyxl
        aba = self.v_aba.get()
        wb = openpyxl.load_workbook(self.path, read_only=True, data_only=True, keep_vba=False)
        ws = wb[aba]
        self._log(f"Pre-visualizacao da aba '{aba}' (primeiras 5 linhas):")
        for i, row in enumerate(ws.iter_rows(max_row=5, values_only=True)):
            self._log(str(row[:15]))
            if i >= 4:
                break
        wb.close()

    def _importar(self):
        import openpyxl
        aba = self.v_aba.get()
        self.log.delete("1.0", "end")
        self._log(f"Importando aba '{aba}'...")
        wb = openpyxl.load_workbook(self.path, read_only=True, data_only=True, keep_vba=False)
        ws = wb[aba]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        header = None
        start = 0
        for i, row in enumerate(rows):
            vals = [str(v).upper().strip() if v else "" for v in row]
            if any(k in vals for k in ["CODIGO", "COD.", "COD.", "COLUNA1", "ITEM"]):
                header = vals
                start = i + 1
                break
        if header is None:
            header = [str(v).upper().strip() if v else f"C{j}" for j, v in enumerate(rows[0])]
            start = 1
        self._log(f"Cabecalho detectado: {header[:12]}")

        def col(h, *nomes):
            for nm in nomes:
                for j, v in enumerate(h):
                    if nm in v:
                        return j
            return None

        i_cod = col(header, "COLUNA1", "COD.", "CODIGO", "ITEM")
        i_ent = col(header, "ENTIDADE")
        i_quad = col(header, "QUADRA")
        i_lote = col(header, "LOTE")
        i_end = col(header, "ENDERECO", "LOGRADOURO")
        i_num = col(header, "NR", "NUMERO")
        i_tipo = col(header, "TIPO")
        i_area = col(header, "AREA")
        i_constr = col(header, "CONSTR")
        i_mat = col(header, "MATRICULA")
        i_local = col(header, "LOCAL")
        i_morador = col(header, "MORADOR")
        i_nct = col(header, "CONTRATO")
        i_tct = col(header, "MODALIDADE")
        i_vig = col(header, "VIGENCIA", "VIGENCIA")
        i_sit = col(header, "SITUACAO", "STATUS")
        i_txoc = col(header, "TX. OCUP.", "TAXA")
        i_pmin = col(header, "PRECO MINIMO", "VALOR MINIMO")

        ok = 0
        skip = 0
        with get_conn() as conn:
            for row in rows[start:]:
                if all(v is None for v in row):
                    continue
                def g(i):
                    return row[i] if i is not None and i < len(row) else None
                cod = str(g(i_cod) or "").strip()
                end = str(g(i_end) or "").strip()
                if not cod or not end or cod in ("None", ""):
                    skip += 1
                    continue
                vig = g(i_vig)
                if hasattr(vig, "strftime"):
                    vig = vig.strftime("%Y-%m-%d")
                elif vig:
                    vig = parse_data(str(vig)[:10])
                d = dict(
                    codigo=cod,
                    entidade=str(g(i_ent) or "").strip() or None,
                    quadra=g(i_quad), lote=g(i_lote),
                    endereco=end,
                    numero=str(g(i_num) or "").strip() or None,
                    tipo=str(g(i_tipo) or "").strip() or None,
                    area_m2=g(i_area), construcao_m2=g(i_constr),
                    matricula=g(i_mat),
                    local_area=str(g(i_local) or "").strip() or None,
                    morador=str(g(i_morador) or "").strip() or None,
                    num_contrato=str(g(i_nct) or "").strip() or None,
                    tipo_contrato=str(g(i_tct) or "").strip() or None,
                    vigencia=vig,
                    sit_contrato=str(g(i_sit) or "").strip() or None,
                    taxa_ocupacao=g(i_txoc),
                    preco_minimo=g(i_pmin),
                )
                try:
                    conn.execute("""INSERT OR IGNORE INTO imoveis
                        (codigo,entidade,quadra,lote,endereco,numero,tipo,area_m2,
                         construcao_m2,matricula,local_area,morador,num_contrato,
                         tipo_contrato,vigencia,sit_contrato,taxa_ocupacao,preco_minimo)
                        VALUES (:codigo,:entidade,:quadra,:lote,:endereco,:numero,:tipo,
                         :area_m2,:construcao_m2,:matricula,:local_area,:morador,
                         :num_contrato,:tipo_contrato,:vigencia,:sit_contrato,
                         :taxa_ocupacao,:preco_minimo)""", d)
                    ok += 1
                except Exception as e:
                    self._log(f"  Erro linha {cod}: {e}")
                    skip += 1
        self._log(f"\nImportacao concluida: {ok} inseridos, {skip} ignorados/erros.")
