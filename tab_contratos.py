import tkinter as tk
from tkinter import ttk, messagebox
from database import get_conn
from widgets import tv_padrao, btn, FormDlg, AZUL, VERMELHO, VERDE, LARANJA, CINZA
from utils import fmt_data, parse_data, STATUS_CONTRATO, MODALIDADES, TIPOS_USO


class TabContratos(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent, padding=12)
        self._build()
        self.carregar()

    def _build(self):
        topo = ttk.Frame(self)
        topo.pack(fill="x", pady=(0, 8))
        ttk.Label(topo, text="Contratos - Cessao Onerosa / Comodato",
                  style="Header.TLabel").pack(side="left")
        btn(topo, "+ Novo", self.novo, "Accent.TButton").pack(side="right", padx=3)
        btn(topo, "Editar", self.editar).pack(side="right", padx=3)
        btn(topo, "Excluir", self.excluir, "Danger.TButton").pack(side="right", padx=3)
        btn(topo, "Importar XLSX", self.importar).pack(side="right", padx=3)

        fil = ttk.Frame(self)
        fil.pack(fill="x", pady=(0, 6))
        ttk.Label(fil, text="Buscar:").pack(side="left")
        self.v_busca = tk.StringVar()
        self.v_busca.trace_add("write", lambda *_: self.carregar())
        ttk.Entry(fil, textvariable=self.v_busca, width=28).pack(side="left", padx=6)
        self.v_status = tk.StringVar(value="TODOS")
        cb = ttk.Combobox(fil, textvariable=self.v_status,
                          values=["TODOS"] + STATUS_CONTRATO, width=12, state="readonly")
        cb.pack(side="left", padx=4)
        cb.bind("<<ComboboxSelected>>", lambda _: self.carregar())
        self.v_modal = tk.StringVar(value="TODOS")
        cb2 = ttk.Combobox(fil, textvariable=self.v_modal,
                           values=["TODOS"] + MODALIDADES, width=12, state="readonly")
        cb2.pack(side="left", padx=4)
        cb2.bind("<<ComboboxSelected>>", lambda _: self.carregar())

        cols = ["Nr Contrato", "Entidade", "Modalidade", "Tipo Uso", "Ass.Dig.",
                "Vigencia", "Status", "Dias/Vencer", "Prorrog. ate", "CPF/CNPJ",
                "Responsavel", "Telefone", "E-mail"]
        largs = [130, 220, 90, 90, 60, 90, 80, 80, 100, 130, 180, 140, 200]
        self.tv, frm = tv_padrao(self, cols, largs, altura=20)
        frm.pack(fill="both", expand=True)
        self.tv.bind("<Double-1>", lambda _: self.editar())
        self.v_total = tk.StringVar()
        ttk.Label(self, textvariable=self.v_total, foreground="#666").pack(anchor="w", pady=2)

    def carregar(self):
        for i in self.tv.get_children():
            self.tv.delete(i)
        busca = f"%{self.v_busca.get()}%"
        status = self.v_status.get()
        modal = self.v_modal.get()
        with get_conn() as conn:
            q = """SELECT * FROM contratos WHERE
                   (num_contrato LIKE ? OR entidade LIKE ? OR responsavel LIKE ?
                    OR cpf_cnpj LIKE ? OR email LIKE ?)"""
            p = [busca] * 5
            if status != "TODOS":
                q += " AND status=?"; p.append(status)
            if modal != "TODOS":
                q += " AND modalidade=?"; p.append(modal)
            q += " ORDER BY data_vigencia"
            rows = conn.execute(q, p).fetchall()
        self.v_total.set(f"{len(rows)} registro(s)")
        import datetime
        for r in rows:
            dias = r["dias_vencer"]
            if dias is None and r["data_vigencia"]:
                from utils import dias_para_vencer
                dias = dias_para_vencer(r["data_vigencia"])
            tag = ""
            if r["status"] == "VENCIDO":
                tag = "vencido"
            elif dias is not None and dias <= 30:
                tag = "alerta"
            elif r["status"] == "VIGENTE":
                tag = "vigente"
            self.tv.insert("", "end", iid=r["id"], values=(
                r["num_contrato"], r["entidade"][:35],
                r["modalidade"] or "", r["tipo_uso"] or "",
                r["assinatura_dig"] or "",
                fmt_data(r["data_vigencia"]),
                r["status"], dias or "",
                fmt_data(r["prorrogavel_ate"]),
                r["cpf_cnpj"] or "", r["responsavel"] or "",
                r["telefone"] or "", r["email"] or ""
            ), tags=(tag,))

    def _sel(self):
        sel = self.tv.selection()
        if not sel:
            messagebox.showwarning("Atencao", "Selecione um contrato.")
            return None
        return int(sel[0])

    def novo(self):
        d = ContratoDlg(self)
        self.wait_window(d)
        self.carregar()

    def editar(self):
        cid = self._sel()
        if cid is None:
            return
        with get_conn() as conn:
            row = conn.execute("SELECT * FROM contratos WHERE id=?", (cid,)).fetchone()
        d = ContratoDlg(self, row)
        self.wait_window(d)
        self.carregar()

    def excluir(self):
        cid = self._sel()
        if cid is None:
            return
        if messagebox.askyesno("Confirmar", "Excluir este contrato?"):
            with get_conn() as conn:
                conn.execute("DELETE FROM contratos WHERE id=?", (cid,))
            self.carregar()

    def importar(self):
        from tkinter.filedialog import askopenfilename
        path = askopenfilename(title="Selecionar planilha",
                               filetypes=[("Excel", "*.xlsx *.xlsm"), ("Todos", "*.*")])
        if not path:
            return
        ImportarContratosDlg(self, path)
        self.carregar()


class ContratoDlg(FormDlg):
    def __init__(self, parent, row=None):
        super().__init__(parent, "Editar Contrato" if row else "Novo Contrato", w=700, h=640)
        self.row = row
        self._build()
        if row:
            self._fill()

    def _build(self):
        c = self.corpo
        for i in range(4):
            c.columnconfigure(i, weight=1)
        self.v_nct, _ = self.campo(c, "Nr Contrato *", 0, 0, cs=2, w=36)
        self.v_modal, _ = self.combo(c, "Modalidade", MODALIDADES, 0, 2)
        self.v_uso, _ = self.combo(c, "Tipo Uso", TIPOS_USO, 0, 3)
        self.v_ent, _ = self.campo(c, "Entidade / Permissionaria *", 1, 0, cs=4, w=70)
        self.v_vig, _ = self.campo(c, "Vigencia (dd/mm/aaaa)", 2, 0)
        self.v_status, _ = self.combo(c, "Status", STATUS_CONTRATO, 2, 1)
        self.v_prorrog, _ = self.campo(c, "Prorrogavel ate (dd/mm/aaaa)", 2, 2)
        self.v_assdig, _ = self.campo(c, "Assinatura Digital", 2, 3)
        self.v_cpf, _ = self.campo(c, "CPF / CNPJ", 3, 0, cs=2, w=36)
        self.v_resp, _ = self.campo(c, "Responsavel / Presidente", 3, 2, cs=2, w=36)
        self.v_end_cor, _ = self.campo(c, "Endereco para correspondencia", 4, 0, cs=4, w=70)
        self.v_tel, _ = self.campo(c, "Telefone(s)", 5, 0, cs=2, w=40)
        self.v_email, _ = self.campo(c, "E-mail(s)", 5, 2, cs=2, w=40)
        self.t_obs = self.texto(c, "Observacoes", 6, 0, cs=4, h=3)
        btn(self.rodape, "Salvar", self._salvar, "Accent.TButton").pack(side="right", padx=8, pady=6)
        btn(self.rodape, "Cancelar", self.destroy).pack(side="right", pady=6)

    def _fill(self):
        r = self.row
        self.v_nct.set(r["num_contrato"])
        self.v_modal.set(r["modalidade"] or "ONEROSO")
        self.v_uso.set(r["tipo_uso"] or "RESIDENCIAL")
        self.v_ent.set(r["entidade"])
        self.v_vig.set(fmt_data(r["data_vigencia"]))
        self.v_status.set(r["status"])
        self.v_prorrog.set(fmt_data(r["prorrogavel_ate"]))
        self.v_assdig.set(r["assinatura_dig"] or "")
        self.v_cpf.set(r["cpf_cnpj"] or "")
        self.v_resp.set(r["responsavel"] or "")
        self.v_end_cor.set(r["end_correspondencia"] or "")
        self.v_tel.set(r["telefone"] or "")
        self.v_email.set(r["email"] or "")
        self.t_obs.insert("1.0", r["observacoes"] or "")

    def _salvar(self):
        nct = self.v_nct.get().strip()
        ent = self.v_ent.get().strip()
        if not nct or not ent:
            messagebox.showerror("Erro", "Nr Contrato e Entidade sao obrigatorios.")
            return
        d = dict(
            num_contrato=nct, entidade=ent,
            modalidade=self.v_modal.get(),
            tipo_uso=self.v_uso.get(),
            data_vigencia=parse_data(self.v_vig.get()),
            status=self.v_status.get(),
            prorrogavel_ate=parse_data(self.v_prorrog.get()),
            assinatura_dig=self.v_assdig.get().strip() or None,
            cpf_cnpj=self.v_cpf.get().strip() or None,
            responsavel=self.v_resp.get().strip() or None,
            end_correspondencia=self.v_end_cor.get().strip() or None,
            telefone=self.v_tel.get().strip() or None,
            email=self.v_email.get().strip() or None,
            observacoes=self.t_obs.get("1.0", "end-1c").strip() or None
        )
        with get_conn() as conn:
            if self.row:
                conn.execute("""UPDATE contratos SET num_contrato=:num_contrato,
                    entidade=:entidade, modalidade=:modalidade, tipo_uso=:tipo_uso,
                    data_vigencia=:data_vigencia, status=:status,
                    prorrogavel_ate=:prorrogavel_ate, assinatura_dig=:assinatura_dig,
                    cpf_cnpj=:cpf_cnpj, responsavel=:responsavel,
                    end_correspondencia=:end_correspondencia, telefone=:telefone,
                    email=:email, observacoes=:observacoes WHERE id=?""",
                    {**d, "?": self.row["id"]})
            else:
                conn.execute("""INSERT INTO contratos (num_contrato,entidade,modalidade,
                    tipo_uso,data_vigencia,status,prorrogavel_ate,assinatura_dig,
                    cpf_cnpj,responsavel,end_correspondencia,telefone,email,observacoes)
                    VALUES (:num_contrato,:entidade,:modalidade,:tipo_uso,:data_vigencia,
                    :status,:prorrogavel_ate,:assinatura_dig,:cpf_cnpj,:responsavel,
                    :end_correspondencia,:telefone,:email,:observacoes)""", d)
        self.destroy()


class ImportarContratosDlg(tk.Toplevel):
    def __init__(self, parent, path):
        super().__init__(parent)
        self.title("Importar Contratos - XLSX")
        self.geometry("680x480")
        self.grab_set()
        self.path = path
        self._build()
        self.after(100, self._carregar_abas)

    def _build(self):
        ttk.Label(self, text=f"Arquivo: {self.path.split('/')[-1]}",
                  font=("Segoe UI", 10)).pack(anchor="w", padx=12, pady=8)
        f = ttk.Frame(self)
        f.pack(fill="x", padx=12)
        ttk.Label(f, text="Aba:").pack(side="left")
        self.v_aba = tk.StringVar()
        self.cb_aba = ttk.Combobox(f, textvariable=self.v_aba, state="readonly", width=35)
        self.cb_aba.pack(side="left", padx=6)
        self.log = tk.Text(self, height=18, font=("Consolas", 9))
        self.log.pack(fill="both", expand=True, padx=12, pady=4)
        rodape = ttk.Frame(self)
        rodape.pack(fill="x", padx=12, pady=6)
        btn(rodape, "Importar", self._importar, "Accent.TButton").pack(side="right", padx=4)
        btn(rodape, "Fechar", self.destroy).pack(side="right")

    def _log(self, t):
        self.log.insert("end", t + "\n"); self.log.see("end")

    def _carregar_abas(self):
        import openpyxl
        wb = openpyxl.load_workbook(self.path, read_only=True, data_only=True, keep_vba=False)
        self.cb_aba["values"] = wb.sheetnames
        if wb.sheetnames:
            self.cb_aba.set("TOT. CONTRATOS" if "TOT. CONTRATOS" in wb.sheetnames else wb.sheetnames[0])
        wb.close()

    def _importar(self):
        import openpyxl
        aba = self.v_aba.get()
        wb = openpyxl.load_workbook(self.path, read_only=True, data_only=True, keep_vba=False)
        ws = wb[aba]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        header = None
        start = 0
        for i, row in enumerate(rows):
            vals = [str(v).upper().strip() if v else "" for v in row]
            if "N CONTRATO" in vals or "CONTRATO" in vals:
                header = vals; start = i + 1; break
        if header is None:
            self._log("Cabecalho nao encontrado."); return

        def ci(h, *nomes):
            for nm in nomes:
                for j, v in enumerate(h):
                    if nm in v: return j
            return None

        i_nct = ci(header, "CONTRATO")
        i_ent = ci(header, "ENTIDADE", "PERMISSIONAR")
        i_modal = ci(header, "MODALIDADE")
        i_uso = ci(header, "TIPO_USO", "COLUNA2")
        i_assdig = ci(header, "ASS. DIG", "ASSINATURA")
        i_vig = ci(header, "VIGENCIA")
        i_status = ci(header, "STATUS")
        i_dias = ci(header, "DIAS")
        i_prorrog = ci(header, "PRORROGAVEL")
        i_cpf = ci(header, "CPF")
        i_resp = ci(header, "PRESIDENTE", "RESPONSAVEL")
        i_end = ci(header, "ENDERECO")
        i_tel = ci(header, "TELEFONE")
        i_email = ci(header, "E-MAIL", "EMAIL")

        ok = skip = 0
        with get_conn() as conn:
            for row in rows[start:]:
                if all(v is None for v in row): continue
                def g(i): return row[i] if i is not None and i < len(row) else None
                nct = str(g(i_nct) or "").strip()
                ent = str(g(i_ent) or "").strip()
                if not nct or nct in ("None", ""):
                    skip += 1; continue
                vig = g(i_vig)
                if hasattr(vig, "strftime"): vig = vig.strftime("%Y-%m-%d")
                prorrog = g(i_prorrog)
                if hasattr(prorrog, "strftime"): prorrog = prorrog.strftime("%Y-%m-%d")
                elif prorrog: prorrog = parse_data(str(prorrog)[:10])
                d = dict(
                    num_contrato=nct, entidade=ent or nct,
                    modalidade=str(g(i_modal) or "ONEROSO").strip(),
                    tipo_uso=str(g(i_uso) or "RESIDENCIAL").strip(),
                    assinatura_dig=str(g(i_assdig) or "").strip() or None,
                    data_vigencia=vig.strftime("%Y-%m-%d") if hasattr(vig, "strftime") else vig,
                    status=str(g(i_status) or "VIGENTE").strip(),
                    dias_vencer=g(i_dias), prorrogavel_ate=prorrog,
                    cpf_cnpj=str(g(i_cpf) or "").strip() or None,
                    responsavel=str(g(i_resp) or "").strip() or None,
                    end_correspondencia=str(g(i_end) or "").strip() or None,
                    telefone=str(g(i_tel) or "").strip() or None,
                    email=str(g(i_email) or "").strip() or None,
                )
                try:
                    conn.execute("""INSERT OR IGNORE INTO contratos
                        (num_contrato,entidade,modalidade,tipo_uso,assinatura_dig,
                         data_vigencia,status,dias_vencer,prorrogavel_ate,cpf_cnpj,
                         responsavel,end_correspondencia,telefone,email)
                        VALUES (:num_contrato,:entidade,:modalidade,:tipo_uso,
                         :assinatura_dig,:data_vigencia,:status,:dias_vencer,
                         :prorrogavel_ate,:cpf_cnpj,:responsavel,:end_correspondencia,
                         :telefone,:email)""", d)
                    ok += 1
                except Exception as e:
                    self._log(f"  Erro {nct}: {e}"); skip += 1
        self._log(f"\nConcluido: {ok} inseridos, {skip} ignorados.")
